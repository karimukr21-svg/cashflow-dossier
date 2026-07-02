#!/usr/bin/env python3
"""Phase B — reconciliation + staging for the Treasury cash-flow tool (Session 3).

Builds on parse_cashflow.py (Session 2 parser core). Per workbook:

  RECONCILE  Σ project sheets  vs  the USD AREA-TOTAL / CONSOLIDATED rollup sheet,
             per line per period. Flow lines (Receipts/Payments) drive the verdict;
             balance lines (opening/ending/accumulated) are stocks that do NOT sum
             linearly across projects, so they are reported as memo, not breaks.
             Each cell classified: tie | rounding | real | missing_in_total |
             missing_in_projects.  => the answer to the Tony-vs-Midas pain.

  STAGE      Write the run as an "open run for review" (LifeDB import-run pattern):
             cf_import_runs (1) + cf_staged_rows (the facts) + cf_recon_breaks (the
             detail) + upsert discovered projects into cf_projects. NEVER touches the
             live canonical tables (cf_actuals/cf_forecasts) or cf_versions — that is
             Push/Publish (Sessions 4-5). Area items fold to project_code='_AREA'
             (is_area_item); JV sheets keep their code, tagged is_jv, stored as-extracted.

Usage:
  python reconcile_stage.py "Qatar_Apr_2026_CashFlow_Updated.xlsx"      # reconcile only
  python reconcile_stage.py --all                                        # all 21, report
  python reconcile_stage.py --all --commit                               # + stage to Supabase
  python reconcile_stage.py --all --commit --reset                       # wipe prior open runs first
  python reconcile_stage.py --report recon_report.md --all               # write markdown report
"""
import openpyxl, glob, os, json, datetime, argparse
from collections import defaultdict, Counter

from parse_cashflow import (SRC, AREA_BY_FILE, Resolver, load_ref, parse_sheet,
                            pick_usd_sheets, classify, strip_currency, tight,
                            DEFAULT_AS_OF, MIN_YEAR, detect_header,
                            is_opening_label, is_ending_label)

HERE = os.path.dirname(os.path.abspath(__file__))

# Cycle this "April 2026 Actual" upload proposes to land in (matches as_of 2026-04-30).
PROPOSED_CYCLE = (2026, 5)
PROPOSED_VERSION = '2026-05-PROJ'   # project-grain version; materialized at Push (S4)

# --- area-total / Σ-projects rollup sheet to reconcile against, by filename -----
#     None => single-sheet / no clean decomposition => verdict 'no_total'.
TARGET_SHEET = {
    'AREA QATAR CASH FLOW ACTUAL APRIL 2026 & FORECAST - MOA.xlsx': 'PROJECTS TOTAL USD',
    'Qatar_Apr_2026_CashFlow_Updated.xlsx': 'CONSOL QAR',   # native QAR consolidation (per Karim)
    'Saudi_Apr_2026_CashFlow.xlsx':         'SAR-CONSOLIDATED',
    'UAE_Apr_2026_CashFlow.xlsx':           'CONSOLIDATED AED',
    'Oman_Apr_2026_CashFlow.xlsx':          'CONSOLIDATED OMR',
    'EPSO_Apr_2026_CashFlow.xlsx':          'CONSOLIDATED-AED',   # native AED, not the USD restatement
    # Jordan: the in-sheet "Currency" labels are SWAPPED (the sheet named "- USD"
    # declares JOD and vice-versa); by the values the JOD-valued sheet is
    # "CONSOLIDATED -AREA JOD". Store native JOD, so the area sheet is that one.
    'Jordan_Apr_2026_CashFlow.xlsx':        'CONSOLIDATED -AREA JOD',
    'Astana_Apr_2026_CashFlow.xlsx':        'CONSOLIDATED',
    'BOTSWANA_Apr_2026_CashFlow.xlsx':      'BOT',   # take only the BOT sheet for now
    'KAZH_Apr_2026_CashFlow.xlsx':          'CONSOLIDATED',
    'Morganti_Apr_2026_CashFlow.xlsx':      'CONSOLIDATED',
    'Ivory Coast_Apr_2026_CashFlow.xlsx':   'IVORY COAST XOF FINAL',   # native XOF, single-area
    'Libya_Apr_2026_CashFlow.xlsx':         'LIBYA CONSOLIDATED',
    'MOZAMBIQUE_Apr_2026_CashFlow.xlsx':    'new Consolidated ',
    'CCC Rwanda_Apr_2026_CashFlow.xlsx':    'RWANDA CONSOLIDATED - CCC ',
    'Rwanda_NBIA_Apr_2026_CashFlow.xlsx':   'RWANDA CONSOLIDATED - CCC share',
    'Algeria_Morocco_Apr_2026_CashFlow.xlsx': 'CONSOLIDATED',
    'Egypt 2026 Cash flow Consolidated - April 2026-2.xlsx': 'CONSOLIDATED',
    'Iraq_Apr_2026_CashFlow.xlsx':          None,
    'Nigeria_Apr_2026_CashFlow.xlsx':       'NIGERIA NGN',   # native NGN, single-area
    'CCUW_Apr_2026_CashFlow.xlsx':          None,
}

# Reconciliation runs in the project sheets' NATIVE currency (the source reality:
# only Qatar publishes per-project USD; UAE=AED, Saudi=SAR, most others local — the
# USD figure exists only at the consolidated level). Σ-projects-vs-area-total is a
# currency-INTERNAL consistency check, so native currency is correct. Converting the
# staged rows to USD is a Push/S4 concern (plan §3.4: derive USD from gacc.fx_rates).
# Detected from each sheet's "Currency XXX" header; this map is a fallback override.
CURRENCY_OVERRIDE = {
    'AREA QATAR CASH FLOW ACTUAL APRIL 2026 & FORECAST - MOA.xlsx': 'USD',
    'Qatar_Apr_2026_CashFlow_Updated.xlsx': 'QAR',   # take the native QAR sheets, not the USD restatement
    'Nigeria_Apr_2026_CashFlow.xlsx': 'USD',
    'Jordan_Apr_2026_CashFlow.xlsx': 'JOD',   # in-sheet currency labels are swapped
    'EPSO_Apr_2026_CashFlow.xlsx': 'AED',     # taking the native AED consolidated sheet
    'Iraq_Apr_2026_CashFlow.xlsx': 'USD',     # single CONSOLIDATED sheet, USD '000
    'CCUW_Apr_2026_CashFlow.xlsx': 'USD',     # no in-sheet currency header; area is USD
    'Nigeria_Apr_2026_CashFlow.xlsx': 'NGN',  # taking the native NGN sheet (header mislabels it USD)
    'Ivory Coast_Apr_2026_CashFlow.xlsx': 'XOF',  # taking the IVORY COAST XOF FINAL sheet
    'Egypt 2026 Cash flow Consolidated - April 2026-2.xlsx': 'USD',  # '000 USD
}
CCY_TOKENS = ('USD', 'SAR', 'AED', 'QAR', 'EUR', 'EGP', 'KZT', 'OMR', 'NGN',
              'XOF', 'JOD', 'MAD', 'BWP', 'RWF', 'GBP', 'TND', 'DZD')

# Source-scale normalization to FULL native currency. The area files are NOT uniform:
# some are denominated in full units (KSA full SAR, Oman full OMR), others in '000 of
# their native currency. The in-sheet 'Currency XXX "000"' label is NOT a reliable
# indicator of the actual value scale — verified by comparison to Tony's master (e.g.
# Qatar/Kazakhstan carry NO '000 label yet are '000; the marker can't be trusted).
# Canonical stores FULL native (so USD derives at read via fx_rates), so a per-file
# factor scales the staged values + balances up to full units. Factor is determined
# per area by Tony-calibration during seeding (default 1.0 = no change), then baked
# here so a re-stage / cloud re-upload stays consistent instead of reverting to '000.
SCALE_OVERRIDE = {
    # pushed + normalized (cf_forecasts already ×1000; config keeps a re-stage consistent)
    'Libya_Apr_2026_CashFlow.xlsx':    1000.0,   # '000 USD -> full USD
    'BOTSWANA_Apr_2026_CashFlow.xlsx': 1000.0,   # '000 USD -> full USD
    'Iraq_Apr_2026_CashFlow.xlsx':     1000.0,   # '000 USD -> full USD
    'Nigeria_Apr_2026_CashFlow.xlsx':  1000.0,   # '000 NGN -> full NGN
    'Jordan_Apr_2026_CashFlow.xlsx':   1000.0,   # '000 JOD -> full JOD
    # staged, verified '000 via Tony-calibration — must RE-STAGE before push so the
    # values land full (SCALE_OVERRIDE applies at stage time, not at push).
    'Qatar_Apr_2026_CashFlow_Updated.xlsx': 1000.0,                    # '000 QAR -> full QAR (native QAR sheets, same '000 scale as the USD ones)
    'AREA QATAR CASH FLOW ACTUAL APRIL 2026 & FORECAST - MOA.xlsx': 1000.0,  # '000 USD
    'UAE_Apr_2026_CashFlow.xlsx':      1000.0,   # '000 AED -> full AED
    'KAZH_Apr_2026_CashFlow.xlsx':     1000.0,   # '000 USD -> full USD
    'Morganti_Apr_2026_CashFlow.xlsx': 1000.0,   # '000 USD -> full USD
    'MOZAMBIQUE_Apr_2026_CashFlow.xlsx': 1000.0, # '000 USD -> full USD
    'CCUW_Apr_2026_CashFlow.xlsx':     1000.0,   # '000 USD -> full USD
    'Ivory Coast_Apr_2026_CashFlow.xlsx': 1000.0, # '000 XOF -> full XOF
    'Egypt 2026 Cash flow Consolidated - April 2026-2.xlsx': 1000.0,  # '000 USD -> full USD
}

# Per-file start-year floor (drop history before this year). Default = global MIN_YEAR.
# Mozambique: the PARP sheet carries an older 2020-2025 project history; the area
# consolidated only starts 2025, so the pre-2025 PARP rows create a spurious year with
# flows but no opening anchor. Floor it at 2025 so the area starts clean with a real
# opening balance.
# Kazakhstan: the file's Dec-2024 ending balance != Jan-2025 opening balance (a manual
# reset, Δ243k — the only discontinuity in the 2020-2026 series). The cash chain is
# DERIVED (anchor = first month's opening, then closing = opening + movements; see
# src/lib/derivedBalances.ts), so a chain spanning that break carries the -243 reset
# forward and offsets 2026 from the file. Floor at 2025 (drop 2024 entirely) so the
# earliest anchor is Jan-2025's clean post-break opening → 2025 + 2026 tie the file.
START_YEAR_OVERRIDE = {
    'MOZAMBIQUE_Apr_2026_CashFlow.xlsx': 2025,
    'KAZH_Apr_2026_CashFlow.xlsx': 2025,
    # Egypt: the ACT-2024 column is a single messy history month (offsetting within-group
    # / non-op vs CONSOLIDATED); the cycle horizon is 2025-2026, so drop 2024 (Karim's call).
    'Egypt 2026 Cash flow Consolidated - April 2026-2.xlsx': 2025,
}

# Per-file end-year ceiling (drop years after this). Default = no cap.
# Qatar: the file carries a PLAN 2027 block; for the APR-2026 cycle the canonical
# store keeps the 2024-2026 actual+forecast horizon only (per Karim), so the 2027
# plan year is dropped at stage time.
END_YEAR_OVERRIDE = {
    'Qatar_Apr_2026_CashFlow_Updated.xlsx': 2026,
    'Egypt 2026 Cash flow Consolidated - April 2026-2.xlsx': 2026,   # drop the 2027/2028 plan years
}

# Sheets that, despite classifying as 'project', are NOT real projects to sum:
# by-project In/Out/NET summaries + sub-rollups that would double-count.
EXCLUDE_SHEETS = {
    'Saudi_Apr_2026_CashFlow.xlsx': {'SAUDI', 'TOTALAREA', 'PROJECTS', 'NEWPROJECTS',
                                     'AREA - 26 TO 28', 'PMV - NET'},
    # UAE: 'UAE' is the top rollup. Also exclude the summary/duplicate books Karim
    # flagged: USD-CONSOLIDATED (the USD restatement of CONSOLIDATED AED, the target),
    # NEW / NEW 3 / LEGACY / NMDCCC (working copies + legacy entities already summed
    # into the AREA sheet), TRANSFER OUTSIDE AREA (counterparty breakdown of a
    # within-group line), and NEW SALES (folded elsewhere) — summing any of them
    # double-counts. Within-area transfers on the AREA sheet are dropped separately
    # (see DROP_WITHIN_AREA) since their legs live on these excluded sheets.
    'UAE_Apr_2026_CashFlow.xlsx': {'UAE', 'USD-CONSOLIDATED', 'NEW', 'NEW 3',
                                   'LEGACY', 'NMDCCC', 'TRANSFER OUTSIDE AREA',
                                   'NEW SALES'},
    # transfer/receipts helper tabs that would double-count the area total
    'Oman_Apr_2026_CashFlow.xlsx': {'TRANSFER WITHIN GROUP OMAN', 'Receipts - Completed'},
    # Rwanda: take ONLY the 'RWANDA CONSOLIDATED - CCC share' sheet (it already
    # combines the RWANDA USD + RWANDA RWF entities into the CCC share, USD '000).
    # Exclude the project sheets so the run falls to single-area on that consolidated
    # sheet (the file also bundles unrelated areas — Botswana/Nigeria/Grenada/Zambia).
    'Rwanda_NBIA_Apr_2026_CashFlow.xlsx': {'RWANDA USD', 'RWANDA RWF', 'IVC',
                                           'BOTSWANA', 'NIGERIA', 'GRENADA', 'ZAMBIA'},
    # Nigeria: take the native NGN sheet (single-area); drop the USD restatement.
    'Nigeria_Apr_2026_CashFlow.xlsx': {'NIGERIA USD'},
    # Ivory Coast: take only 'IVORY COAST XOF FINAL'; drop the USD restatements and
    # the unrelated areas the file bundles (Mozam/Rwanda/Botswana/Grenada/Zambia).
    'Ivory Coast_Apr_2026_CashFlow.xlsx': {'CONSOLIDATED', 'IVORY COAST USD FINAL',
        'IVORY COAST XOF', 'MOZAM', 'RWANDA', 'BOTSWANA', 'GRENADA', 'ZAMBIA', 'NEW SALES'},
    # Botswana: take only the BOT sheet for now; the file bundles other areas.
    'BOTSWANA_Apr_2026_CashFlow.xlsx': {'CONSOLIDATED', 'BOT (2)', 'IVC', 'RWANDA',
        'BOTSWANA', 'NIGERIA', 'GRENADA', 'ZAMBIA', 'NEW SALES'},
    # CC (UE): the area's complete cash flow IS the 'CASH FLOW REPORTS' statement,
    # which already carries the within-group net (PAYMENTS - OUTSIDE AREA, row 45).
    # The 'TRANSFERS OUTSIDE AREA' sheet is a counterparty breakdown of that same
    # within-group line — summing it double-counts the transfers (and its TOTAL
    # doesn't even match the statement's row 45). Drop it; the main statement is
    # authoritative and is what the area's balance walk uses. ('NEW SALES' is kept:
    # the main statement's New Sales section is zero, so that sheet is the only
    # source of new-sales flow, not a duplicate.)
    'CCUW_Apr_2026_CashFlow.xlsx': {'TRANSFERS OUTSIDE AREA'},
}

# Sheets whose name contains 'JV' but which are NOT equity-accounted JVs to exclude —
# they're CCC's PROPORTIONAL share (already share-adjusted, e.g. 'PM JV 54%') and the
# area's CONSOLIDATED rollup sums them in. Force them to be treated as normal
# consolidated entities (summed + reconciled), overriding the is_jv_code heuristic
# (which otherwise drops any 'JV'-named sheet, correct for UAE/Saudi's 100% JV books).
JV_INCLUDE_OVERRIDE = {
    # Morganti CONSOLIDATED = PM JV 54% + FM 100% + MGSA 70% (verified to the cent);
    # PM JV 54% is the 54% share, part of the consolidation — sum it.
    'Morganti_Apr_2026_CashFlow.xlsx': {'PM JV 54%'},
    # Egypt: 'Central JV CCCE-CCCEgypt' is a real 50%-share project leaf (its 'JV'-in-name
    # would otherwise drop it); it's part of the CONSOLIDATED consolidation, sum it.
    'Egypt 2026 Cash flow Consolidated - April 2026-2.xlsx': {'Central JV CCCE-CCCEgypt'},
}

# Per-file, per-sheet ownership SHARE. Some areas' project LEAF sheets carry the
# project's 100% cash flow, and the area CONSOLIDATED applies CCC's ownership share
# (e.g. a 50%/60% JV). To stage project-grain that ties CONSOLIDATED, scale each such
# leaf's values by its share at ingest time (before the reconcile). Sheets not listed
# default to 1.0 (no scaling). Distinct from JV_INCLUDE_OVERRIDE, whose share is already
# baked into the sheet values (e.g. Morganti 'PM JV 54%'). Verified per group against
# each sub-rollup (Σ leaves × share == rollup, to the cent).
SHEET_SHARE = {
    'Egypt 2026 Cash flow Consolidated - April 2026-2.xlsx': {
        # 2026 JV
        'Ras Elhekma': 0.5, 'Central JV CCCE-CCCEgypt': 0.5, '2': 0.6,
        'A2': 0.5, 'A3': 0.5, 'A4': 0.5, 'A5': 0.5, 'A6': 0.5, 'A7': 0.5, 'A8': 0.5, 'A9': 0.5,
        # Ongoing JV
        'Alamein': 0.5, 'Marassi': 0.5, 'Marassi MEP': 0.5, 'Madinaty CP08': 0.5,
        'Madinaty CP07': 0.5, 'Madinaty CP05': 0.5, 'Arkan': 0.5, 'Arkan Towers': 0.5,
        'Madinaty CP05 MEP': 0.5, 'FS Luxor MEP': 0.5, 'CCC EGP on': 0.6, 'CCC Egypt 25': 0.6,
        # Legacy JV (Zafarana is 100%, not a JV share — per Karim)
        'Helwan': 0.5, 'Cairo West': 0.5, 'Mivida': 0.5, 'Nile Plaza': 0.5,
        'Madinaty CP03': 0.5, 'CFC': 0.5, 'CFC MEP': 0.5,
        # Time-varying FLOW shares: CCC Egypt leg + CCCEgypt area book Jan-2025 flows at
        # 45%, Feb-2025+ at 60% (they have Jan-2025 activity that steps). CCC EGP on / CCC
        # Egypt 25 stay flat 60% — their Jan flows are already booked at 60% in CONSOLIDATED
        # (stepping them broke the tie), even though CCC EGP on's balance is part of the
        # Feb-2025 share-restatement top-up (a balance-only item; see EGYPT_OPENING_ADJ).
        'CCC  Egypt leg': [(2025, 1, 0.45), (2025, 2, 0.60)],
        'CCCEgypt area': [(2025, 1, 0.45), (2025, 2, 0.60)],
        'CCC EGP on': 0.6,
        'CCC Egypt 25': 0.6,
    },
}

# Drop WITHIN-AREA transfer lines (both legs) for these files. Within-area transfers
# are internal to the area and net to zero at the area level — the consolidated confirms
# it (equal-and-opposite each month). UAE routes its within-area transfers through the
# AREA sheet, which is a rollup that already sums the excluded NEW/LEGACY entities; once
# those are excluded the AREA sheet's within-area leg is orphaned (its counterparty lived
# on the excluded sheets), so within-area no longer nets to zero and inflates the tie in
# 4 months of 2025. Dropping within-area on BOTH sides reconciles the area on its EXTERNAL
# transfers (outside-area + treasury) only — which is what actually moves area cash.
DROP_WITHIN_AREA = {'UAE_Apr_2026_CashFlow.xlsx'}

# Disable the per-sheet within-group sign flip for these files. The flip (parse_sheet)
# negates a sheet's within-group payments when their sum is positive, to catch files
# that store transfer-OUT detail as gross magnitudes (e.g. CCUW). But when the area is
# signed throughout — proven by the CONSOL rollup carrying within-area PAY as negative
# that nets to zero — a lone signed-positive within-group payment (a transfer reversal)
# must stay positive to keep that netting. Qatar's LCB has exactly one within-area
# payment (Feb-2025 +90); flipping it to -90 breaks the area's Feb-2025 within-area net
# (a -180 phantom). Every other Qatar sheet already stores within-group signed, so
# disabling the flip for the file changes only LCB and matches CONSOL.
NO_WG_SIGN_FLIP = {'Qatar_Apr_2026_CashFlow_Updated.xlsx'}
WITHIN_AREA_CODES = {'wg_recpt_within_area', 'wg_pay_within_area'}

# Force-include sheets the classifier drops as rollups but which carry real area data
# that must be summed. UAE's 'AREA' sheet is named like a rollup (so pick_usd_sheets
# skips it), but it is the ONLY source of the area's external within-group transfers
# (outside-area + treasury) — without it the area's within-group collapses. It does not
# duplicate project operations, so summing it is correct (its orphaned within-area legs
# are dropped via DROP_WITHIN_AREA).
FORCE_INCLUDE_SHEETS = {
    'UAE_Apr_2026_CashFlow.xlsx': {'AREA'},
}

# Explicit summed-set override: when present, REPLACES pick_usd_sheets — the run sums
# exactly these sheets (by stripped name). For multi-currency files where the auto
# picker would grab the USD restatement but we want the NATIVE-currency sheets (e.g.
# Qatar carries a parallel USD + QAR book of every project; take the QAR set + the
# QAR consolidated as target). EXCLUDE_SHEETS / FORCE_INCLUDE_SHEETS still apply after.
SHEET_INCLUDE_OVERRIDE = {
    'Qatar_Apr_2026_CashFlow_Updated.xlsx': {
        'BHP', 'BOP', 'BSNP', 'DHP', 'DIAP', 'DIAR', 'DPCT', 'GSF', 'JSPP', 'LCB',
        'LMH2', 'LR2', 'MAT', 'MDP1', 'MDP4', 'MPAC', 'NDAB', 'NFE', 'QFIS', 'QIN4',
        'QRDB', 'QRMS', 'QRMS MAINT', 'RGX', 'RAHP', 'RLP', 'UMRP', 'WMFF', 'WMRP',
        'NFE BLDG', 'NFE EPC.1', 'QAFCO7', 'NFS', 'RLPP', 'NGL-5', 'HIAR', 'ITCA',
        'NFW', 'CCG', 'TCC', 'CCG D', 'MORG',
    },
    # Egypt: sum the project LEAF sheets only (JV leaves scaled by SHEET_SHARE), tie
    # CONSOLIDATED. Ignore the rollups (CONSOLIDATED / Consolidated 2026|Ongoing|Legacy /
    # the CCC & JV sub-rollups / AREA / OFFSHORE / Summery / NEW SALES / helper sheets).
    # OFFSHORE (CCIC/CCCE/CCC RE/CCCEgypt area) is a separate holding/treasury layer, not
    # in CONSOLIDATED — excluded (Karim's call: projects only for now).
    # Egypt: Karim's reviewed 44-sheet set (empty project sheets removed). Leaf projects
    # + the 4 offshore/holding entities + OFFSHORE; JV/offshore leaves scaled by SHEET_SHARE.
    'Egypt 2026 Cash flow Consolidated - April 2026-2.xlsx': {
        'Central JV CCCE-CCCEgypt', 'CCIC', 'I-City MEP', 'Central Offshore Main', 'Cairo West',
        'City Gate 3D', 'Madinaty CP03', 'City Gate 2D', 'Alamein', 'Derna', 'I-City',
        'CCC  Egypt leg', 'CFC MEP', 'Arkan Towers', 'Madinaty CP05', 'CFC', 'CCC EGP on',
        'FS Luxor MEP', 'Madinaty CP08', 'Ras Elhekma', 'Shepheard CCCE', 'WE CCCE',
        'City Gate 3D Compo MEP1', 'CCCEgypt area', 'City Gate 2C', 'Nile Plaza',
        'Madinaty CP05 MEP', 'Madinaty CP07', 'WE', 'SQ1 CCCE1', 'City Gate 3D Infra,Villas Comp',
        'CCCE', 'City Gate 2C MEP', 'OFFSHORE', 'Arkan', 'CCC RE', 'Zafarana', 'Mivida',
        'Marassi MEP', 'Marassi', 'City Gate Infra', 'CCC Egypt 25', 'City Gate MEP', 'Helwan',
    },
}

def _resolve_share(spec, year, month):
    """Resolve a SHEET_SHARE spec to a factor for (year, month). spec is either a flat
    float, or a list of (year, month, share) breakpoints sorted ascending — the share of
    the latest breakpoint whose (year, month) <= the row's applies (steps forward)."""
    if isinstance(spec, (int, float)):
        return spec
    val = 1.0
    for (y, m, s) in spec:
        if (y, m) <= (year, month):
            val = s
        else:
            break
    return val


# Per-file, per-sheet line-code DROP. Some project leaves carry a line the area's
# CONSOLIDATED rollup does NOT consolidate; to match the file, drop just that line from
# that sheet (keeping the sheet's other lines). Egypt: CONSOLIDATED omits Ras Elhekma's
# overdraft financing (bf_od) — draws + repayments are on its sheet but not in CONSOL, so
# drop them to match (documented in the data-issues report as a file consolidation gap).
DROP_SHEET_LINES = {
    'Egypt 2026 Cash flow Consolidated - April 2026-2.xlsx': {
        'Ras Elhekma': {'bf_recpt_od', 'bf_pay_od'},
    },
}

# Single-entity areas: no per-project decomposition exists, only an area cash flow.
# Stage that sheet under project_code='_AREA' (verdict 'single_area' — nothing to tie).
MAIN_SHEET = {
    'Iraq_Apr_2026_CashFlow.xlsx': 'CONSOLIDATED',
    'EPSO_Apr_2026_CashFlow.xlsx': 'CONSOLIDATED-USD',
    'Jordan_Apr_2026_CashFlow.xlsx': 'CONSOLIDATED - USD',
    'Astana_Apr_2026_CashFlow.xlsx': 'CONSOLIDATED',
    # no-total areas: the sheet that carries the area's BALANCE AT START anchor.
    'CCUW_Apr_2026_CashFlow.xlsx': 'CASH FLOW REPORTS',
}


# Currency word-forms -> canonical token. Header cells say "Currency US$" /
# "(Currency US Dollar)" / "QR '000" / "Dirhams" — normalise them to a CCY token.
# Checked before the bare CCY_TOKENS so 'US$'/'DOLLAR' resolve to USD, not a stray
# letter match. A bare '$' near 'CURRENCY' is treated as USD as a last resort.
CCY_SYNONYMS = (
    ('US$', 'USD'), ('US $', 'USD'), ('U.S', 'USD'), ('US DOLLAR', 'USD'),
    ('USDOLLAR', 'USD'), ('DOLLAR', 'USD'), ('EURO', 'EUR'), ('DIRHAM', 'AED'),
    ('RIYAL', 'SAR'), ('TENGE', 'KZT'), ('DINAR', 'JOD'),
)


def _ccy_from_header(up):
    """Map an uppercased header cell to a currency token, or None."""
    for word, tok in CCY_SYNONYMS:
        if word in up:
            return tok
    for tok in CCY_TOKENS:
        if tok in up:
            return tok
    if '$' in up:
        return 'USD'
    return None


def detect_currency(wb, sheets):
    """Majority currency read from a 'Currency XXX' / "'000" header line (rows 1-3).
    Recognises word-forms (US$, US Dollar, Dirham, QR…) as well as the bare tokens.
    Scans the given project sheets first; if none vote, falls back to scanning EVERY
    sheet — the currency is often declared only on the consolidated rollup (e.g.
    Kazakhstan: 'CASH FLOW With Claims (Currency US$)' lives on CONSOLIDATED)."""
    def vote_over(names):
        votes = Counter()
        for n in names:
            try:
                ws = wb[n]
            except KeyError:
                continue
            for row in ws.iter_rows(min_row=1, max_row=3, max_col=6, values_only=True):
                for v in row:
                    if not isinstance(v, str):
                        continue
                    up = v.upper()
                    if 'CURRENCY' in up or "'000" in up or '"000"' in up or "000)" in up:
                        tok = _ccy_from_header(up)
                        if tok:
                            votes[tok] += 1
        return votes
    votes = vote_over(sheets[:8])
    if not votes:
        votes = vote_over(wb.sheetnames)
    return votes.most_common(1)[0][0] if votes else '?'

def detect_target_sheet(wb, native_ccy):
    """Auto-detect the area-total / consolidated rollup sheet (the recon target,
    and the sheet basis-B reads the area balance from) WITHOUT a per-file map, so
    a new cycle's file (different name) still self-identifies its rollup.

    A candidate is any sheet whose name says CONSOLIDATED / CONSOL / PROJECTS
    TOTAL. Score prefers the area's native currency, de-prioritises prev/legacy
    restatements. Reproduces the tuned TARGET_SHEET map for 19/21 known files;
    the two ambiguous ones keep an explicit override in TARGET_SHEET. Returns the
    sheet name or None."""
    cands = []
    for n in wb.sheetnames:
        up = n.upper()
        base, cur = strip_currency(n)
        if not ('CONSOLIDAT' in up or 'CONSOL' in up or 'PROJECTS TOTAL' in up):
            continue
        score = 0
        if native_ccy and native_ccy != '?':
            if cur == native_ccy:
                score += 100
            if native_ccy in up:
                score += 100
        if cur == 'USD' or 'USD' in up:
            score += 10
        if 'PROJECTS TOTAL' in up:
            score += 5
        if 'PREV' in up or 'LEGACY' in up or 'OLD' in up:
            score -= 50
        cands.append((score, len(n), n))
    if not cands:
        return None
    cands.sort(key=lambda x: (-x[0], x[1]))
    return cands[0][2]


def extract_rollup_balances(ws, as_of):
    """Basis B: lift the area's OWN maintained opening/ending cash balance from the
    rollup sheet, verbatim. Project sheets either open at 0 (e.g. KSA, movement-only)
    or carry their own balances (e.g. Qatar) — so the authoritative area balance is
    the rollup's BALANCE AT START / BALANCE AT END row, not a derivation.

    First-row-wins per kind: the rollup commonly prints the native liquid-funds
    ending followed by a USD ('000) restatement; taking the first match keeps the
    native-currency series and drops the duplicate (which otherwise summed into
    garbage).

    Also lifts the area's debt STOCKS verbatim — accumulated loans and overdraft
    balances (the running closing-debt rows below BALANCE AT END). These are stocks,
    not flows, so summing them across project sheets double-counts; the rollup is the
    authoritative source. Returns {'opening','ending','accum_loans','accum_od'},
    each {'YYYY-MM': v}."""
    rows = list(ws.iter_rows(min_row=1, max_row=90, max_col=120, values_only=True))
    hdr = detect_header(rows, as_of)
    empty = {'opening': {}, 'ending': {}, 'accum_loans': {}, 'accum_od': {},
             'accum_loans_open': None, 'accum_od_open': None}
    if not hdr:
        return empty
    lc = hdr['label_col']
    periods = hdr['periods']
    first_ci = min(periods.keys()) if periods else None
    opening, ending, accum_loans, accum_od = {}, {}, {}, {}
    accum_loans_open = accum_od_open = None
    got_open = got_end = got_accl = got_accod = False

    def series(r):
        out = {}
        for ci, (y, m, _k) in periods.items():
            if ci < len(r) and isinstance(r[ci], (int, float)):
                out[f'{y:04d}-{m:02d}'] = round(float(r[ci]), 4)
        return out

    # the debt-stock OPENING (before the first month) sits one row BELOW the stock
    # label, in the column just left of the first month (Karim's layout note). The
    # accumulated balance then rolls forward from there by the bank-financing movement.
    def open_below(i):
        if first_ci is None or i + 1 >= len(rows):
            return None
        nb = rows[i + 1]
        v = nb[first_ci - 1] if 0 <= first_ci - 1 < len(nb) else None
        return round(float(v), 4) if isinstance(v, (int, float)) else None

    body = list(enumerate(rows))[hdr['header_row'] + 1:]
    for i, r in body:
        label = r[lc] if lc < len(r) and isinstance(r[lc], str) else None
        if not label or not label.strip():
            continue
        lt = tight(label)
        if not got_open and is_opening_label(lt):
            opening = series(r); got_open = True
        elif not got_end and is_ending_label(lt):
            ending = series(r); got_end = True
        # debt stocks sit BELOW the ending balance (the closing-debt block)
        elif got_end and not got_accl and ('ACCOUMULATEDLOAN' in lt or 'ACCUMULATEDLOAN' in lt):
            accum_loans = series(r); accum_loans_open = open_below(i); got_accl = True
        elif got_end and not got_accod and lt.startswith('OVERDRAFT'):
            accum_od = series(r); accum_od_open = open_below(i); got_accod = True
    return {'opening': opening, 'ending': ending,
            'accum_loans': accum_loans, 'accum_od': accum_od,
            'accum_loans_open': accum_loans_open, 'accum_od_open': accum_od_open}


def section_key(category, nature):
    """Map a line's (category, nature) to the grid's section bucket — the same
    grouping the cash-flow grid uses, so the rollup's OWN section totals line up
    1:1 with Σ-projects for the variance comparison."""
    if category in ('Operation', 'Claims', 'New Sales') and nature == 'Receipts':
        return 'oper_rec'
    if category in ('Operation', 'New Sales') and nature == 'Payments':
        return 'oper_pay'
    return {'Interest': 'interest', 'Non Operational': 'nonop',
            'Within Group': 'wg', 'Bank Financing': 'bank'}.get(category)


def sections_from_agg(agg, ref):
    """Roll a parsed sheet's (line_code, year, month)->value map up into the grid's
    section totals per month. Used on the rollup so the file's stated movements can
    be diffed against Σ-projects."""
    cat = {L['line_code']: L['category'] for L in ref['lines']}
    nat = {L['line_code']: L['nature'] for L in ref['lines']}
    out = defaultdict(lambda: defaultdict(float))
    for (lc, y, m), v in agg.items():
        if y < MIN_YEAR:
            continue
        k = section_key(cat.get(lc), nat.get(lc))
        if k:
            out[k][f'{y:04d}-{m:02d}'] += round(v, 4)
    return {k: dict(v) for k, v in out.items()}


# Area-item sheet tokens (fold to project_code='_AREA', is_area_item=true) -------
AREA_ITEM_TOKENS = ('PMV', 'CAMP', 'RMPT', 'OVERHEAD', 'RECOVER', 'AREAOH', 'MOEST')

# Lines that are STOCKS (not flows) — reported as memo in reconciliation ----------
BALANCE_LINES = {'opening_balance', 'ending_balance', 'accum_loans', 'accum_od'}


def is_area_item(code):
    t = tight(code)
    return any(t == tok or t.startswith(tok) for tok in AREA_ITEM_TOKENS)


def is_jv_code(code):
    return 'JV' in tight(code)


def classify_diff(diff, total, sum_proj):
    """rounding | real | missing_in_total | missing_in_projects."""
    a = abs(diff)
    tol = max(1.0, 0.01 * abs(total))          # 1 (=$1k) or 1% of the total cell
    if a <= tol:
        return 'rounding'
    if abs(total) <= tol and abs(sum_proj) > tol:
        return 'missing_in_total'
    if abs(sum_proj) <= tol and abs(total) > tol:
        return 'missing_in_projects'
    return 'real'


def parse_target(wb, area, sheet, resolver, as_of, wg_sign_flip=True):
    rows, meta = parse_sheet(wb[sheet], area, sheet, resolver, as_of,
                             wg_sign_flip=wg_sign_flip)
    agg = defaultdict(float)
    if rows:
        for r in rows:
            agg[(r['line_code'], r['year'], r['month'])] += r['value']
    return agg, (meta or {})


def reconcile_workbook(path, resolver, as_of):
    fn = os.path.basename(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return _reconcile_with_wb(wb, fn, resolver, as_of)


def reconcile_workbook_bytes(file_bytes, fn, resolver, as_of):
    """Serverless entry point: reconcile an uploaded workbook from raw bytes.
    (Added for the api/cf-stage.py Vercel function — keeps a single engine.)"""
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    return _reconcile_with_wb(wb, fn, resolver, as_of)


def _reconcile_with_wb(wb, fn, resolver, as_of):
    area = AREA_BY_FILE.get(fn, fn.split('_')[0])
    wg_flip = fn not in NO_WG_SIGN_FLIP   # per-file within-group sign-flip toggle
    all_sheets = list(wb.sheetnames)
    if fn in SHEET_INCLUDE_OVERRIDE:
        _inc = {s.strip() for s in SHEET_INCLUDE_OVERRIDE[fn]}
        sel = [n for n in all_sheets if n.strip() in _inc]
    else:
        sel = pick_usd_sheets(wb)
    excl = EXCLUDE_SHEETS.get(fn, set())
    sel = [s for s in sel if s.strip() not in {e.strip() for e in excl}]
    # force-include rollup-named sheets that carry real area data (see FORCE_INCLUDE_SHEETS)
    _excl_strip = {e.strip() for e in excl}
    for finc in FORCE_INCLUDE_SHEETS.get(fn, set()):
        if finc in all_sheets and finc not in sel and finc.strip() not in _excl_strip:
            sel.append(finc)
    currency = CURRENCY_OVERRIDE.get(fn) or detect_currency(wb, sel)

    # parse every selected project sheet, assign project_code + flags
    proj_rows = []            # normalized rows with project_code resolved
    projects = {}             # project_code -> {sheet, is_area_item, is_jv, n}
    unmatched_labels = {}     # label -> {count, locations: ["Sheet!C12", ...]}

    def _merge_unmatched(src):
        for lab, info in (src or {}).items():
            # tolerate the old {label: count} shape just in case
            cnt = info['count'] if isinstance(info, dict) else info
            cells = info.get('cells', []) if isinstance(info, dict) else []
            agg = unmatched_labels.setdefault(lab, {'count': 0, 'locations': []})
            agg['count'] += cnt
            for cell in cells:
                if cell not in agg['locations'] and len(agg['locations']) < 8:
                    agg['locations'].append(cell)
    nat = {L['line_code']: L['nature'] for L in resolver_lines(resolver)}

    # Per-sheet metadata for the interactive include/ignore toggle. EVERY parseable
    # sheet is staged (each row stamped with its source sheet) so a sheet can be
    # moved between Included and Ignored after the fact. default_included reproduces
    # today's summed set (the `sel` sheets, non-JV); everything else stages
    # default-excluded and is opt-in in the UI — this is how a real area component
    # the classifier hides as a 'rollup' (e.g. KSA 'AREA') can be pulled back in.
    sheet_meta = {}        # sheet -> {sheet, code, role, is_jv, name, default_included}
    sheet_unmatched = {}   # sheet -> {label: {count, cells}} (for ALL staged sheets)

    def _role_of(code, n, area_item, jv):
        if jv:
            return 'jv'
        if area_item or code == '_AREA':
            return 'area_item'
        if classify(n)[0] == 'rollup':
            return 'rollup'
        return 'assigned' if resolver.gacc.get(tight(code)) else 'unassigned'

    def _ingest(n, default_inc, dest):
        """Parse sheet n, stamp project_code + sheet on every row into `dest`, and
        record sheet_meta. Only default-included sheets feed `projects`/unmatched
        (the recon basis stays exactly today's behaviour)."""
        rows, meta = parse_sheet(wb[n], area, n, resolver, as_of, wg_sign_flip=wg_flip)
        if rows is None:
            return
        raw_code = meta['project_code']
        area_item = is_area_item(raw_code)
        # 'JV'-named sheet → excluded as equity-accounted, UNLESS this file marks it a
        # proportional-share entity the rollup consolidates (e.g. Morganti 'PM JV 54%').
        jv = is_jv_code(raw_code) and n not in JV_INCLUDE_OVERRIDE.get(fn, set())
        # area items (non-JV) fold to the _AREA bucket
        code = '_AREA' if (area_item and not jv) else raw_code
        # per-sheet ownership share (see SHEET_SHARE): leaf carries 100%, scale to CCC's
        # share so Σ-projects ties the share-based CONSOLIDATED. Applied before reconcile.
        # A share may be a flat float or a period-stepped list of (year, month, share).
        share_spec = SHEET_SHARE.get(fn, {}).get(n, 1.0)
        drop_lines = DROP_SHEET_LINES.get(fn, {}).get(n, set())
        for r in rows:
            if r['line_code'] in drop_lines:
                continue   # line the CONSOLIDATED rollup doesn't consolidate (see DROP_SHEET_LINES)
            r = dict(r); r['project_code'] = code; r['sheet'] = n
            share = _resolve_share(share_spec, r['year'], r['month'])
            if share != 1.0:
                r['value'] = round(r['value'] * share, 4)
            dest.append(r)
        gi = resolver.gacc.get(tight(code))
        sheet_meta[n] = {'sheet': n, 'code': code,
                         'role': _role_of(code, n, area_item, jv), 'is_jv': jv,
                         'name': (gi.get('abbreviation') if gi else None),
                         'default_included': bool(default_inc) and not jv}
        if meta.get('unmatched'):
            sheet_unmatched[n] = meta['unmatched']   # per-sheet, regardless of default
        if default_inc:
            p = projects.setdefault(code, {'sheets': [], 'is_area_item': False,
                                           'is_jv': False, 'n': 0})
            p['sheets'].append(n)
            p['is_area_item'] = p['is_area_item'] or area_item
            p['is_jv'] = p['is_jv'] or jv
            p['n'] += len(rows)
            _merge_unmatched(meta['unmatched'])

    for n in sel:
        _ingest(n, True, proj_rows)

    # basis B: the area's own maintained opening/ending balance + its own section
    # totals, read verbatim from the rollup (set below once the target/main sheet is
    # known). rollup_sections lets the staging UI diff Σ-projects against the file.
    rollup_balances = {'opening': {}, 'ending': {}}
    rollup_sections = {}
    ref_lines = {'lines': resolver_lines(resolver)}

    # single-entity areas: no per-project decomposition — stage the area's own
    # cash-flow sheet under _AREA (nothing to reconcile against).
    single_area = False
    if not proj_rows:
        fb = (TARGET_SHEET.get(fn) or MAIN_SHEET.get(fn)
              or detect_target_sheet(wb, currency))
        if fb and fb in wb.sheetnames:
            rows, meta = parse_sheet(wb[fb], area, fb, resolver, as_of, wg_sign_flip=wg_flip)
            if rows:
                for r in rows:
                    r = dict(r); r['project_code'] = '_AREA'; r['sheet'] = fb
                    proj_rows.append(r)
                projects['_AREA'] = {'sheets': [fb], 'is_area_item': True,
                                     'is_jv': False, 'n': len(rows)}
                sheet_meta[fb] = {'sheet': fb, 'code': '_AREA', 'role': 'area_item',
                                  'is_jv': False, 'name': None, 'default_included': True}
                if meta.get('unmatched'):
                    sheet_unmatched[fb] = meta['unmatched']
                _merge_unmatched(meta['unmatched'])
                rollup_balances = extract_rollup_balances(wb[fb], as_of)
                # file == our data here, so file sections = our sections (variance 0)
                fb_agg = defaultdict(float)
                for r in rows:
                    fb_agg[(r['line_code'], r['year'], r['month'])] += r['value']
                rollup_sections = sections_from_agg(fb_agg, ref_lines)
                single_area = True

    # Per-file within-area drop (see DROP_WITHIN_AREA) — internal transfers that net to
    # zero at the area level; remove both legs so the area ties on external transfers only.
    if fn in DROP_WITHIN_AREA:
        proj_rows = [r for r in proj_rows if r['line_code'] not in WITHIN_AREA_CODES]

    # aggregate project rows to canonical grain (sum within-run dups, e.g. _AREA)
    agg = defaultdict(float)                    # (proj, line, y, m) -> value (ALL, for staging)
    for r in proj_rows:
        agg[(r['project_code'], r['line_code'], r['year'], r['month'])] += r['value']
    # reconciliation sum EXCLUDES JV projects: area consolidated sheets are
    # equity-accounted (JV cash flows are not line-consolidated) — plan §3.8.
    # JVs are still STAGED as-extracted (is_jv); they just don't enter the tie.
    jv_codes = {c for c, p in projects.items() if p['is_jv']}
    flowsum = defaultdict(float)                # (line, y, m) -> Σ over NON-JV projects
    for (pc, lc, y, m), v in agg.items():
        if pc in jv_codes:
            continue
        flowsum[(lc, y, m)] += v

    # reconcile against the target rollup sheet (skip for single-entity areas)
    MATERIAL = {'real', 'missing_in_total', 'missing_in_projects'}
    # explicit per-file override wins (incl. an explicit None); otherwise auto-detect
    # the consolidated rollup so a new cycle's filename self-identifies its target.
    target = TARGET_SHEET[fn] if fn in TARGET_SHEET else detect_target_sheet(wb, currency)
    breaks = []
    recon_status = 'single_area' if single_area else 'no_total'
    target_used = None
    # Balance anchor for no-total areas: read opening/ending from a designated MAIN
    # sheet even without a reconcile target (e.g. CCUW has no consolidated total, but
    # CASH FLOW REPORTS carries the area's BALANCE AT START — the running-balance anchor).
    if not single_area and not (target and target in wb.sheetnames):
        ms = MAIN_SHEET.get(fn)
        if ms and ms in wb.sheetnames:
            rollup_balances = extract_rollup_balances(wb[ms], as_of)
    if target and target in wb.sheetnames and not single_area:
        tgt, _ = parse_target(wb, area, target, resolver, as_of, wg_sign_flip=wg_flip)
        if fn in DROP_WITHIN_AREA:
            tgt = {k: v for k, v in tgt.items() if k[0] not in WITHIN_AREA_CODES}
        target_used = target
        rollup_balances = extract_rollup_balances(wb[target], as_of)
        rollup_sections = sections_from_agg(tgt, ref_lines)
        keys = set(flowsum) | set(tgt)
        material = 0
        for k in sorted(keys):
            lc, y, m = k
            if y < MIN_YEAR:
                continue
            sp = round(flowsum.get(k, 0.0), 4)
            tv = round(tgt.get(k, 0.0), 4)
            diff = round(sp - tv, 4)
            is_bal = lc in BALANCE_LINES
            if abs(diff) < 0.5:
                continue                         # exact tie (incl. balances), skip noise
            cls = 'balance_memo' if is_bal else classify_diff(diff, tv, sp)
            if cls in MATERIAL:
                material += 1
            breaks.append({'line_code': lc, 'nature': nat.get(lc), 'year': y, 'month': m,
                           'sum_projects': sp, 'area_total': tv, 'diff': diff,
                           'classification': cls})
        recon_status = 'tie' if material == 0 else 'break'
    # No independent area total to reconcile against (a single consolidated/area
    # sheet, or an area with no rollup): the INCLUDED sheets ARE the area total, so
    # set the file side = our own section sums. Without this the compare-to-file
    # toggle diffs against an empty file side and shows every cell as a phantom
    # variance (the values are correct, there's just nothing to compare to).
    if not rollup_sections and flowsum:
        rollup_sections = sections_from_agg(flowsum, ref_lines)
    flow_breaks = [b for b in breaks if b['classification'] in MATERIAL]
    max_abs = max((abs(b['diff']) for b in flow_breaks), default=None)
    cat = {L['line_code']: L['category'] for L in resolver_lines(resolver)}
    by_cat = Counter(cat.get(b['line_code'], '?') for b in flow_breaks)
    n_brk_act = sum(1 for b in flow_breaks
                    if datetime.date(b['year'], b['month'], 1) <= as_of)
    n_brk_fc = len(flow_breaks) - n_brk_act

    # Stage every OTHER parseable sheet as default-excluded so it can be toggled
    # in later (the recon above is untouched — it only ever saw the `sel` sheets).
    extra_rows = []
    for n in all_sheets:
        if n in sheet_meta or n == target:
            continue
        _ingest(n, False, extra_rows)
    wb.close()

    # Staging aggregate keeps the SHEET in the key, so area items that fold to the
    # same _AREA bucket (PMV/CAMP/RMPT/AREA) stay individually toggleable.
    if fn in DROP_WITHIN_AREA:
        extra_rows = [r for r in extra_rows if r['line_code'] not in WITHIN_AREA_CODES]
    stage_agg = defaultdict(float)
    for r in proj_rows + extra_rows:
        stage_agg[(r['project_code'], r['sheet'], r['line_code'], r['year'], r['month'])] += r['value']
    staged = []
    n_act = n_fc = 0
    for (pc, sheet, lc, y, m), v in stage_agg.items():
        if y < MIN_YEAR:
            continue
        kind = 'actual' if datetime.date(y, m, 1) <= as_of else 'forecast'
        if kind == 'actual':
            n_act += 1
        else:
            n_fc += 1
        staged.append({'kind': kind, 'area': area, 'project_code': pc, 'sheet': sheet,
                       'line_code': lc, 'year': y, 'month': m, 'value': round(v, 4)})

    # Normalize source '000 -> full native (see SCALE_OVERRIDE). Default 1.0 = no-op.
    # Scales the staged values AND the rollup balance/section series together, so the
    # compare-to-file tie and the balance tiles stay internally consistent.
    scale = SCALE_OVERRIDE.get(fn, 1.0)
    if scale != 1.0:
        for r in staged:
            r['value'] = round(r['value'] * scale, 4)
        for bk in ('opening', 'ending', 'accum_loans', 'accum_od'):
            if isinstance(rollup_balances.get(bk), dict):
                rollup_balances[bk] = {k: round(v * scale, 4)
                                       for k, v in rollup_balances[bk].items()}
        for bk in ('accum_loans_open', 'accum_od_open'):
            if isinstance(rollup_balances.get(bk), (int, float)):
                rollup_balances[bk] = round(rollup_balances[bk] * scale, 4)
        rollup_sections = {sk: {k: round(v * scale, 4) for k, v in d.items()}
                           for sk, d in rollup_sections.items()}

    # Per-file start-year floor (see START_YEAR_OVERRIDE). Drops older history that
    # has no opening anchor, so the area starts clean. Recompute the A/F counts after.
    start_year = START_YEAR_OVERRIDE.get(fn)
    if start_year:
        staged = [r for r in staged if r['year'] >= start_year]
        for bk in ('opening', 'ending', 'accum_loans', 'accum_od'):
            if isinstance(rollup_balances.get(bk), dict):
                rollup_balances[bk] = {k: v for k, v in rollup_balances[bk].items()
                                       if int(k[:4]) >= start_year}
        rollup_sections = {sk: {k: v for k, v in d.items() if int(k[:4]) >= start_year}
                           for sk, d in rollup_sections.items()}
        n_act = sum(1 for r in staged if r['kind'] == 'actual')
        n_fc = sum(1 for r in staged if r['kind'] == 'forecast')

    # Per-file end-year ceiling (see END_YEAR_OVERRIDE). Drops out-of-horizon plan
    # years, symmetric to the start-year floor. Applied to staged rows + the rollup
    # balances/sections so the recon stays like-for-like. Recompute A/F counts after.
    end_year = END_YEAR_OVERRIDE.get(fn)
    if end_year:
        staged = [r for r in staged if r['year'] <= end_year]
        for bk in ('opening', 'ending', 'accum_loans', 'accum_od'):
            if isinstance(rollup_balances.get(bk), dict):
                rollup_balances[bk] = {k: v for k, v in rollup_balances[bk].items()
                                       if int(k[:4]) <= end_year}
        rollup_sections = {sk: {k: v for k, v in d.items() if int(k[:4]) <= end_year}
                           for sk, d in rollup_sections.items()}
        n_act = sum(1 for r in staged if r['kind'] == 'actual')
        n_fc = sum(1 for r in staged if r['kind'] == 'forecast')

    summed_set = set(sel)
    # Per summed sheet: each is a project or an area item. Split into ASSIGNED
    # (the project mapped onto the canonical gacc registry) vs UNASSIGNED (summed
    # as a project but not recognised — needs attention) vs AREA ITEMS (the _AREA
    # bucket: PMV/CAMP/RMPT/overheads). Everything else = ignored (rollups/junk).
    assigned, unassigned, area_items = [], [], []
    for code, p in projects.items():
        for sheet in p['sheets']:
            if code == '_AREA' or p['is_area_item']:
                area_items.append({'sheet': sheet, 'code': code})
            else:
                gi = resolver.gacc.get(tight(code))
                if gi:
                    assigned.append({'sheet': sheet, 'code': code,
                                     'name': gi.get('abbreviation') or code,
                                     'is_jv': p['is_jv']})
                else:
                    unassigned.append({'sheet': sheet, 'code': code, 'is_jv': p['is_jv']})
    # Flat per-sheet list (workbook order) driving the include/ignore toggle, plus
    # the default-included set the run is seeded with. compare_target is the rollup
    # the statement reconciles to (the file side) — shown, never toggled.
    sheets_list = [sheet_meta[n] for n in all_sheets if n in sheet_meta]
    included_sheets = [m['sheet'] for m in sheets_list if m['default_included']]
    sheet_classification = {
        'target': target_used,
        'compare_target': target,
        'sheets': sheets_list,
        'summed': sorted(sel),
        'assigned': sorted(assigned, key=lambda x: x['sheet']),
        'unassigned': sorted(unassigned, key=lambda x: x['sheet']),
        'area_items': sorted(area_items, key=lambda x: x['sheet']),
        'ignored': sorted(n for n in all_sheets
                          if n not in summed_set and n != target_used),
    }

    return {
        'area': area, 'file': fn, 'currency': currency,
        'as_of': as_of, 'n_sheets': len(sel), 'projects': projects,
        'n_projects': len(projects), 'unmatched_labels': dict(unmatched_labels),
        'rollup_balances': rollup_balances, 'rollup_sections': rollup_sections,
        'sheet_classification': sheet_classification, 'included_sheets': included_sheets,
        'unmatched_by_sheet': sheet_unmatched,
        'recon_status': recon_status, 'recon_target': target_used,
        'n_real_breaks': len(flow_breaks),
        'n_rounding': sum(1 for b in breaks if b['classification'] == 'rounding'),
        'max_abs_diff': max_abs, 'breaks': breaks,
        'breaks_by_category': dict(by_cat),
        'n_breaks_actual': n_brk_act, 'n_breaks_forecast': n_brk_fc,
        'staged': staged, 'n_actual': n_act, 'n_forecast': n_fc,
    }


def resolver_lines(resolver):
    return load_ref()['lines']


# ---- report -----------------------------------------------------------------
def fmt(res, full=False):
    L = []
    v = res['recon_status']
    badge = {'tie': 'TIE ✓', 'break': 'BREAK', 'no_total': 'no area total',
             'single_area': 'single-entity area (no decomposition)',
             'unknown': 'unknown'}.get(v, v)
    L.append(f"=== {res['area']} ({res['file']}) — {badge} ===")
    L.append(f"  currency={res['currency']}  sheets={res['n_sheets']}  projects={res['n_projects']}"
             f"  staged: {res['n_actual']}A/{res['n_forecast']}F"
             f"  unmatched_labels={len(res['unmatched_labels'])}")
    if res['recon_target']:
        L.append(f"  target='{res['recon_target']}'  real_breaks={res['n_real_breaks']}"
                 f"  ({res['n_breaks_actual']} actual / {res['n_breaks_forecast']} forecast)"
                 f"  rounding={res['n_rounding']}  max|Δ|={res['max_abs_diff']}")
        if res['breaks_by_category']:
            L.append("  breaks by category: " + ", ".join(
                f"{c}={n}" for c, n in sorted(res['breaks_by_category'].items(),
                                              key=lambda x: -x[1])))
    ai = [c for c, p in res['projects'].items() if p['is_area_item']]
    jv = [c for c, p in res['projects'].items() if p['is_jv']]
    if ai: L.append(f"  area-items(_AREA): {', '.join(ai)}")
    if jv: L.append(f"  JV: {', '.join(jv)}")
    if res['unmatched_labels']:
        L.append("  UNMATCHED: " + ", ".join(
            f"{k!r}x{(v['count'] if isinstance(v, dict) else v)}"
            for k, v in res['unmatched_labels'].items()))
    rb = [b for b in res['breaks'] if b['classification'] == 'real']
    if rb:
        L.append(f"  -- real flow breaks ({len(rb)}) --")
        for b in sorted(rb, key=lambda x: -abs(x['diff']))[:(999 if full else 10)]:
            L.append(f"     {b['line_code']:22} {b['year']}-{b['month']:02d}  "
                     f"Σproj={b['sum_projects']:12.1f} total={b['area_total']:12.1f} Δ={b['diff']:11.1f}")
    return "\n".join(L)


# ---- staging (REST writes) --------------------------------------------------
def _proj_rows(res, ref):
    """The cf_projects rows for a parsed result (skips the _AREA bucket)."""
    gid = ref.get('gacc_id_index', {})
    gidx = ref.get('gacc_index', {})
    rows = []
    for code, p in res['projects'].items():
        if code == '_AREA':
            continue
        rows.append({'project_code': code, 'area': res['area'],
                     'display_name': (gidx.get(tight(code), {}) or {}).get('abbreviation') or code,
                     'gacc_project_id': gid.get(tight(code)),
                     'is_area_item': p['is_area_item'], 'is_jv': p['is_jv']})
    return rows


def _run_summary(res):
    """The cf_import_runs.recon_summary jsonb for a parsed result."""
    return {
        'target_sheet': res['recon_target'], 'currency': res['currency'],
        'projects': {c: {'sheets': p['sheets'], 'is_area_item': p['is_area_item'],
                         'is_jv': p['is_jv'], 'rows': p['n']}
                     for c, p in res['projects'].items()},
        'unmatched_labels': res['unmatched_labels'],
        'unmatched_by_sheet': res.get('unmatched_by_sheet', {}),
        'breaks_by_class': dict(Counter(b['classification'] for b in res['breaks'])),
        'breaks_by_category': res['breaks_by_category'],
        'breaks_actual': res['n_breaks_actual'], 'breaks_forecast': res['n_breaks_forecast'],
        # basis B: the area's own opening/ending balance, read verbatim from the
        # rollup (run-scoped metadata — NOT a staged fact, so it never reaches the
        # canonical store on push). cf_run_review surfaces it as the balance tiles.
        'rollup_balances': res.get('rollup_balances', {'opening': {}, 'ending': {}}),
        'rollup_sections': res.get('rollup_sections', {}),
        'sheet_classification': res.get('sheet_classification'),
    }


def stage(res, ref, created_by='reconcile_stage.py'):
    import db
    # 1. upsert discovered projects into cf_projects
    projrows = _proj_rows(res, ref)
    if projrows:
        db.insert('cf_projects', projrows, upsert=True, on_conflict='project_code')

    # 2. cf_import_runs (the open run)
    summary = _run_summary(res)
    # Staged UNASSIGNED: no cycle/as_of/proposed_version until the run is pushed
    # into a chosen version. n_forecast_rows carries the full staged total (the
    # actuals/forecast split is cycle-dependent, resolved at push).
    run = db.insert_returning('cf_import_runs', [{
        'area': res['area'], 'source_file': res['file'], 'as_of_date': None,
        'cycle_year': None, 'cycle_month': None,
        'proposed_version': None, 'currency': res['currency'],
        'status': 'open', 'n_sheets': res['n_sheets'], 'n_projects': res['n_projects'],
        'n_projects_new': len(projrows), 'n_actual_rows': 0,
        'n_forecast_rows': res['n_actual'] + res['n_forecast'], 'n_unmatched_labels': len(res['unmatched_labels']),
        'recon_target_sheet': res['recon_target'], 'recon_status': res['recon_status'],
        'recon_n_breaks': res['n_real_breaks'], 'recon_n_rounding': res['n_rounding'],
        'recon_max_abs_diff': res['max_abs_diff'], 'recon_summary': summary,
        'included_sheets': res['included_sheets'],
        'created_by': created_by,
    }])[0]
    run_id = run['run_id']

    # 3. cf_staged_rows
    srows = [{'run_id': run_id, **r} for r in res['staged']]
    db.insert('cf_staged_rows', srows)

    # 4. cf_recon_breaks (skip exact ties; keep classified breaks + balance memo)
    brows = [{'run_id': run_id, **b} for b in res['breaks']]
    if brows:
        db.insert('cf_recon_breaks', brows)
    return run_id


def restage_in_place(run_id, res, ref):
    """Re-parse result -> overwrite the SAME run's staged rows + summary, KEEPING the
    user's Included/Ignored sheet selection. Used by the Re-apply Mappings action: a
    line alias added in the UI only takes effect when the file is parsed again."""
    import db
    # preserve the saved sheet selection (drop any sheet that no longer exists)
    cur = db.select('cf_import_runs',
                    {'select': 'included_sheets', 'run_id': f'eq.{run_id}'})
    saved = (cur[0].get('included_sheets') if cur else None) or []
    new_sheets = {s['sheet'] for s in (res.get('sheet_classification') or {}).get('sheets', [])}
    kept = [s for s in saved if s in new_sheets]
    included = kept if kept else res['included_sheets']

    projrows = _proj_rows(res, ref)
    if projrows:
        db.insert('cf_projects', projrows, upsert=True, on_conflict='project_code')

    # swap this run's staged rows + breaks (run_id unchanged)
    db.delete('cf_staged_rows', {'run_id': f'eq.{run_id}'})
    db.delete('cf_recon_breaks', {'run_id': f'eq.{run_id}'})
    db.insert('cf_staged_rows', [{'run_id': run_id, **r} for r in res['staged']])
    brows = [{'run_id': run_id, **b} for b in res['breaks']]
    if brows:
        db.insert('cf_recon_breaks', brows)

    db.update('cf_import_runs', {'run_id': f'eq.{run_id}'}, {
        'source_file': res['file'], 'currency': res['currency'],
        'n_sheets': res['n_sheets'], 'n_projects': res['n_projects'],
        'n_projects_new': len(projrows),
        'n_forecast_rows': res['n_actual'] + res['n_forecast'],
        'n_unmatched_labels': len(res['unmatched_labels']),
        'recon_target_sheet': res['recon_target'], 'recon_status': res['recon_status'],
        'recon_n_breaks': res['n_real_breaks'], 'recon_n_rounding': res['n_rounding'],
        'recon_max_abs_diff': res['max_abs_diff'], 'recon_summary': _run_summary(res),
        'included_sheets': included,
    })
    return run_id


# ---- CLI --------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('file', nargs='?')
    ap.add_argument('--all', action='store_true')
    ap.add_argument('--as-of', default=None)
    ap.add_argument('--commit', action='store_true', help='write staging to Supabase')
    ap.add_argument('--reset', action='store_true', help='delete prior OPEN runs first')
    ap.add_argument('--report', default=None, help='write a markdown report')
    ap.add_argument('--full', action='store_true', help='list all real breaks')
    args = ap.parse_args()
    as_of = datetime.date.fromisoformat(args.as_of) if args.as_of else DEFAULT_AS_OF
    ref = load_ref()
    resolver = Resolver(ref)

    if args.reset and args.commit:
        import db
        db.delete('cf_import_runs', {'status': 'eq.open'})
        print("reset: deleted prior open runs (+ cascade)\n")

    files = sorted(glob.glob(SRC + '*.xlsx')) if args.all else \
            [args.file if os.path.exists(args.file) else SRC + args.file]
    out_lines, grand = [], []
    for p in files:
        res = reconcile_workbook(p, resolver, as_of)
        block = fmt(res, full=args.full)
        print(block); print()
        out_lines.append(block)
        if args.commit:
            rid = stage(res, ref)
            print(f"   staged run {rid}\n")
        grand.append((res['area'], res['recon_status'], res['n_real_breaks'],
                      res['n_actual'] + res['n_forecast'], len(res['unmatched_labels']),
                      res['n_breaks_actual'], res['n_breaks_forecast'],
                      res['currency'], res['recon_target']))

    print("==== SUMMARY ====")
    for area, st, rb, n, ul, ba, bf, ccy, tgt in grand:
        print(f"  {area:14} {st:11} breaks={rb:<4} staged_rows={n:<6} unmatched={ul}")
    if args.report:
        with open(args.report, 'w') as f:
            f.write("# Phase B — Reconciliation report (Session 3)\n\n")
            f.write("_Σ project sheets vs the area-total / consolidated rollup, per line "
                    "per period, in each area's native currency. JV sheets excluded from "
                    "the tie (equity-accounted, not line-consolidated) but staged "
                    "as-extracted. Balance lines (opening/ending/accum) are stocks → memo, "
                    "not breaks. Cutover 2026-04-30._\n\n")
            f.write("## Verdict summary\n\n")
            f.write("| Area | Verdict | Breaks (act/fcst) | Staged | Currency | Target sheet |\n")
            f.write("|------|---------|-------------------|--------|----------|-------------|\n")
            for area, st, rb, n, ul, ba, bf, ccy, tgt in grand:
                f.write(f"| {area} | {st} | {rb} ({ba}/{bf}) | {n} | {ccy} | "
                        f"{tgt or '—'} |\n")
            f.write("\n## Per-area detail\n\n")
            f.write("\n\n".join(out_lines))
        print("\nWrote", args.report)


if __name__ == '__main__':
    main()
