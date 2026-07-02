#!/usr/bin/env python3
"""Label-anchored parser for Treasury project-level cash-flow workbooks.

Phase B core (Session 2). Per workbook:
  - classify project vs rollup/check/local-currency sheets
  - pick the USD sheet (name suffix, or USD column detection)
  - detect the label column (drifts B<->C) and the section column to its left
  - handle all three period axes: datetime / month-name / year-grouped
    (year-grouped reads the year-block header AND the month sub-row)
  - resolve each label -> line_code using SECTION + DIRECTION context
    (the same label, e.g. OTHERS / LTR / OVERDRAFT, is Receipts in one block and
    Payments in another -- so resolution is contextual, not a flat dict)
  - resolve sheet name -> project_code, best-effort link to gacc.projects
  - apply the ACTUAL/PLAN cutover to tag each row actual vs forecast
  - ignore pre-2024 periods

Output: normalized rows (area, project_code, line_code, year, month, value, kind)
plus an unmatched-labels and unmatched-projects report. Does NOT write the DB
(that is Session 3).

Usage:
  python parse_cashflow.py "Qatar_Apr_2026_CashFlow_Updated.xlsx" [--as-of 2026-04-30] [--out qatar_rows.json] [-v]
  python parse_cashflow.py --all          # sweep every workbook, summary only
"""
import openpyxl, glob, os, json, re, datetime, argparse, sys
from collections import defaultdict, Counter
from openpyxl.utils import get_column_letter

SRC = '/Users/karimhakawati/Library/CloudStorage/OneDrive-CCC/GACC - 05. TREASURY/Areas April 2026 Actual/'
HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_AS_OF = datetime.date(2026, 4, 30)   # "April 2026 Actual" cutover
MIN_YEAR = 2024                               # ignore pre-2024 periods

MONTHS = {m: i+1 for i, m in enumerate(
    ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC'])}

# Filename -> canonical area (matches public.cf_actuals.area vocabulary).
AREA_BY_FILE = {
    'AREA QATAR CASH FLOW ACTUAL APRIL 2026 & FORECAST - MOA.xlsx': 'Qatar',
    'Qatar_Apr_2026_CashFlow_Updated.xlsx': 'Qatar',
    'Algeria_Morocco_Apr_2026_CashFlow.xlsx': 'Algeria',
    'Astana_Apr_2026_CashFlow.xlsx': 'Kazakhstan',
    'BOTSWANA_Apr_2026_CashFlow.xlsx': 'Botswana',
    'CCC Rwanda_Apr_2026_CashFlow.xlsx': 'Rwanda',
    'CCUW_Apr_2026_CashFlow.xlsx': 'CC (UE)',
    'EPSO_Apr_2026_CashFlow.xlsx': 'EPSO',
    'Egypt 2026 Cash flow Consolidated - April 2026-2.xlsx': 'Egypt',
    'Iraq_Apr_2026_CashFlow.xlsx': 'Iraq',
    'Ivory Coast_Apr_2026_CashFlow.xlsx': 'Ivory Coast',
    'Jordan_Apr_2026_CashFlow.xlsx': 'Jordan',
    'KAZH_Apr_2026_CashFlow.xlsx': 'Kazakhstan',
    'Libya_Apr_2026_CashFlow.xlsx': 'Lybia',
    'MOZAMBIQUE_Apr_2026_CashFlow.xlsx': 'Mozambique',
    'Morganti_Apr_2026_CashFlow.xlsx': 'Morganti',
    'Nigeria_Apr_2026_CashFlow.xlsx': 'Nigeria',
    'Oman_Apr_2026_CashFlow.xlsx': 'Oman',
    'Rwanda_NBIA_Apr_2026_CashFlow.xlsx': 'Rwanda',
    'Saudi_Apr_2026_CashFlow.xlsx': 'KSA',
    'UAE_Apr_2026_CashFlow.xlsx': 'UAE',
}

# ---- reference data ---------------------------------------------------------
def load_ref():
    with open(os.path.join(HERE, 'ref_data.json')) as f:
        ref = json.load(f)
    # Live-merge the canonical line catalog + aliases from the DB so edits made in
    # the Treasury workspace (inline line-mapping in staging) take effect on the
    # NEXT upload — no ref_data.json regen/redeploy. The DB is the source of truth;
    # this snapshot is only the offline fallback. Any failure (e.g. local CLI with
    # no service key) silently keeps the bundled snapshot. gacc_index is left as-is.
    try:
        import db
        lines = db.select('cf_lines',
            {'select': 'line_code,nature,category,description,is_active', 'order': 'sort_order'})
        aliases = db.select('cf_line_aliases',
            {'select': 'alias_description,alias_nature,alias_category,line_code'})
        if lines:
            ref['lines'] = [
                {'line_code': r['line_code'], 'nature': r['nature'], 'category': r['category'],
                 'description': r['description'], 'is_active': r.get('is_active', True)}
                for r in lines
            ]
        if aliases is not None:
            ref['aliases'] = [
                {'alias': r['alias_description'], 'nature': r['alias_nature'],
                 'category': r['alias_category'], 'line_code': r['line_code']}
                for r in aliases if r.get('alias_description')
            ]
    except Exception:
        pass  # bundled ref_data.json is the fallback
    return ref

def tight(s):
    return re.sub(r'[^A-Z0-9]', '', (s or '').upper())

class Resolver:
    """Contextual label -> line_code resolver (section + direction aware)."""
    def __init__(self, ref):
        self.by_full = {}      # (tight_desc, nature, category) -> code
        self.by_nat = {}       # (tight_desc, nature) -> code  (only if unique)
        self.gacc = ref['gacc_index']
        # line_code -> category / nature, for sign normalisation in parse_sheet
        self.cat_by_code = {L['line_code']: L['category'] for L in ref['lines']}
        self.nat_by_code = {L['line_code']: L['nature'] for L in ref['lines']}
        nat_keys = defaultdict(set)
        for L in ref['lines']:
            if not L['is_active']:
                continue
            t = tight(L['description'])
            self.by_full[(t, L['nature'], L['category'])] = L['line_code']
            nat_keys[(t, L['nature'])].add(L['line_code'])
        for A in ref['aliases']:
            t = tight(A['alias'])
            self.by_full[(t, A['nature'], A['category'])] = A['line_code']
            nat_keys[(t, A['nature'])].add(A['line_code'])
        for k, codes in nat_keys.items():
            if len(codes) == 1:
                self.by_nat[k] = next(iter(codes))

    def lookup(self, label, nature, category):
        t = tight(label)
        if (t, nature, category) in self.by_full:
            return self.by_full[(t, nature, category)]
        # fall back across category when (label, nature) is unambiguous
        if (t, nature) in self.by_nat:
            return self.by_nat[(t, nature)]
        return None

# ---- sheet classification ---------------------------------------------------
CURRENCY_SUFFIXES = ('USD', 'QAR', 'AED', 'SAR', 'EUR', 'KZT', 'OMR', 'EGP')
ROLLUP_TOKENS = ['CONSOLIDAT', 'CONSOL', 'AREA TOTAL', 'AREA OFF', 'AREA WLL',
                 'AREA TCC', 'AREA QAR', 'PROJECTS TOTAL', 'NEW CCG', 'LEGACY',
                 'CHECK', 'BREAKDOWN', 'PLAN ', 'PREV', 'SUMMERY', 'SUMMARY',
                 "JV'S 100", 'PJTS', 'TRANSFER OUTSIDE', 'RECEIPTS FROM',
                 'PAYMENTS TO', 'PAYMENT TERMS', 'PROJECT VALUE', 'SHEET',
                 # stale / reference tabs kept in the workbook for history — not a
                 # live project, so never sum them (e.g. 'KSA_for records').
                 'FOR RECORD', 'RECORDS', 'BACKUP', 'OBSOLETE', 'DO NOT USE']
AREA_ROLLUP_EXACT = {'AREA', 'NEW', 'CCG', 'CCG D', 'TCC', 'MORG', 'CONSOLIDATED',
                     'OFFSHORE', 'CCIC', 'CCCE', 'CCC RE', 'PJTS', 'A2','A3','A4',
                     'A5','A6','A7','A8','A9','AA','CC','DD','EE','FF','GG','D','E',
                     'F','2','YY','YYY','WE'}

def strip_currency(name):
    """'RLPP USD ' -> ('RLPP','USD'); 'BHP' -> ('BHP', None)."""
    n = name.strip()
    up = n.upper()
    for cur in CURRENCY_SUFFIXES:
        if up.endswith(' ' + cur):
            return n[:-(len(cur))].strip(), cur
        if up == cur:
            return '', cur
    return n, None

def classify(name):
    """Return 'project' | 'rollup' | 'check' | 'consolidated_usd'."""
    base, cur = strip_currency(name)
    up = name.upper().strip()
    bup = base.upper().strip()
    if 'CHECK' in up:
        return 'check', cur
    for tok in ROLLUP_TOKENS:
        if tok in up:
            return 'rollup', cur
    if bup in AREA_ROLLUP_EXACT or up in AREA_ROLLUP_EXACT:
        return 'rollup', cur
    return 'project', cur

# ---- header / axis detection ------------------------------------------------
def is_date(v):
    return isinstance(v, (datetime.datetime, datetime.date))

def month_num(v):
    if not isinstance(v, str):
        return None
    return MONTHS.get(v.strip().upper()[:3])

def month_year(v):
    """(month, explicit_year|None) from a header token: 'JAN' -> (1,None);
    "JAN'26" -> (1,2026); 'JAN-2025' -> (1,2025). The 2-digit suffix lets a
    bare-month axis that switches years mid-row (e.g. EPSO: JAN..DEC = prior
    year, then JAN'26.. = cutover year) be anchored to the real year."""
    if not isinstance(v, str):
        return (None, None)
    s = v.strip().upper()
    mn = MONTHS.get(s[:3])
    if mn is None:
        return (None, None)
    m = re.search(r"(?:'|\s|-|/)(\d{4}|\d{2})\b", s[3:])
    yr = None
    if m:
        g = m.group(1)
        yr = int(g) if len(g) == 4 else 2000 + int(g)
    return (mn, yr)

def yearblock_tokens(v):
    if not isinstance(v, str):
        return None
    u = v.upper()
    if re.search(r'\b(ACT|ACTUAL|FORECAST|FCAST|PLAN|BUDGET)\b', u):
        yrs = [int(y) for y in re.findall(r'(20\d\d)', u)]
        kind = 'forecast' if re.search(r'FORECAST|FCAST|PLAN|BUDGET', u) else 'actual'
        return {'kind': kind, 'years': yrs}
    return None

def find_label_col(row, max_col=None):
    """Column index of the DESCRIPTION / line-label cell in a header row.
    If max_col is given, only consider columns strictly left of it — the label
    column is always left of the period axis (guards against bespoke blocks,
    e.g. Saudi's quarterly 'expected cash' template sitting to the RIGHT)."""
    best = None
    for ci, v in enumerate(row):
        if max_col is not None and ci >= max_col:
            break
        if isinstance(v, str) and 'DESCRIPTION' in v.upper():
            return ci
        if isinstance(v, str) and len(v.strip()) > 2 and month_num(v) is None \
                and not yearblock_tokens(v) and best is None:
            best = ci
    return best if best is not None else 1


def label_col_left_of(rows, header_ri, first_period_col):
    """Find the label column left of the period axis, scanning the header row and
    up to 3 rows above it for a DESCRIPTION cell; else the nearest text column."""
    for ri in range(max(0, header_ri - 3), header_ri + 1):
        for ci, v in enumerate(rows[ri]):
            if ci >= first_period_col:
                break
            if isinstance(v, str) and 'DESCRIPTION' in v.upper():
                return ci
    # fall back to the rightmost text column left of the period axis
    best = None
    for ri in range(max(0, header_ri - 3), header_ri + 1):
        for ci, v in enumerate(rows[ri]):
            if ci >= first_period_col:
                break
            if isinstance(v, str) and len(v.strip()) > 2 and month_num(v) is None \
                    and not yearblock_tokens(v):
                best = ci
    return best if best is not None else min(1, max(0, first_period_col - 1))

def detect_header(rows, as_of):
    """Return dict: {mode, label_col, section_col, periods:{col:(year,month,kind)}}
    or None if no period axis found."""
    # --- datetime axis (also folds in month-name cells in the SAME row) ---
    # Some sheets mix datetime columns with bare/text month tokens that carry the
    # prior-year actuals (e.g. Nigeria: JAN..SEP then 2025-10.. ; CCUW: "JAN 25 - A"
    # then 2026-05..). The datetimes anchor the year; bare months are filled around
    # them with a two-pass walk so those early-2025 actuals are not dropped.
    for ri, row in enumerate(rows[:10]):
        dates = [(ci, v) for ci, v in enumerate(row) if is_date(v)]
        if len(dates) >= 3:
            # collect every period-ish column: (col, month, explicit_year|None)
            cols = []
            for ci, v in enumerate(row):
                if is_date(v):
                    d = v.date() if isinstance(v, datetime.datetime) else v
                    cols.append((ci, d.month, d.year))
                elif month_num(v) is not None:
                    mn, ey = month_year(v)
                    cols.append((ci, mn, ey))
            cols.sort()
            # forward fill years (snap to explicit, JAN rollover for bare)
            yr = {}
            cur = None; prevm = None
            for ci, mn, ey in cols:
                if ey is not None:
                    cur = ey
                elif cur is not None and prevm is not None and mn < prevm:
                    cur += 1
                yr[ci] = cur; prevm = mn
            # backward fill any leading bare months (before the first anchor)
            cur = None; nextm = None
            for ci, mn, ey in reversed(cols):
                if yr[ci] is not None:
                    cur = yr[ci]; nextm = mn; continue
                if ey is not None:
                    cur = ey
                elif cur is not None and nextm is not None and mn > nextm:
                    cur -= 1
                yr[ci] = cur; nextm = mn
            periods = {}
            for ci, mn, _ey in cols:
                y = yr[ci]
                if y is None:
                    continue
                kind = 'actual' if datetime.date(y, mn, 1) <= as_of else 'forecast'
                periods[ci] = (y, mn, kind)
            first_period_col = min(periods) if periods else min(ci for ci, _ in dates)
            lc = label_col_left_of(rows, ri, first_period_col)
            return {'mode': 'datetime', 'header_row': ri, 'label_col': lc,
                    'section_col': max(0, lc - 1), 'periods': periods}
    # --- year-grouped axis ---
    for ri, row in enumerate(rows[:10]):
        blocks = [(ci, yb) for ci, v in enumerate(row)
                  for yb in [yearblock_tokens(v)] if yb]
        if not blocks:
            continue
        # month sub-row within the next 2 rows
        for mr in range(ri + 1, min(ri + 3, len(rows))):
            months = [(ci, *month_year(v)) for ci, v in enumerate(rows[mr])
                      if month_num(v) is not None]   # [(col, month, explicit_year|None)]
            if len(months) >= 3:
                lc = find_label_col(row)
                if find_label_col(rows[mr]) is not None and not is_date(rows[mr][lc] if lc < len(rows[mr]) else None):
                    pass
                periods = _year_grouped_periods(blocks, months, as_of)
                return {'mode': 'year-grouped', 'header_row': mr, 'label_col': lc,
                        'section_col': max(0, lc - 1), 'periods': periods}
    # --- bare month-name axis (single year, inferred) ---
    for ri, row in enumerate(rows[:10]):
        months = [(ci, *month_year(v)) for ci, v in enumerate(row)
                  if month_num(v) is not None]   # [(col, month, explicit_year|None)]
        if len(months) >= 3:
            lc = find_label_col(row)
            # base year: a 4-digit token in the header rows, else as_of year
            yr = None
            for r in rows[:ri + 1]:
                for v in r:
                    if isinstance(v, str):
                        m = re.search(r'(20\d\d)', v)
                        if m:
                            yr = int(m.group(1))
            base = yr or as_of.year
            # provisional forward-walk (JAN rollover -> +1)
            prov = {}
            prev = 0
            y = base
            for ci, mn, _ey in months:
                if mn < prev:
                    y += 1
                prev = mn
                prov[ci] = y
            # if any column carries an EXPLICIT year (e.g. JAN'26), snap the whole
            # sequence to it — fixes axes that start in a prior year then switch.
            delta = 0
            for ci, mn, ey in months:
                if ey is not None:
                    delta = ey - prov[ci]
                    break
            periods = {}
            for ci, mn, _ey in months:
                y = prov[ci] + delta
                d = datetime.date(y, mn, 1)
                periods[ci] = (y, mn, 'actual' if d <= as_of else 'forecast')
            return {'mode': 'month-name', 'header_row': ri, 'label_col': lc,
                    'section_col': max(0, lc - 1), 'periods': periods, 'year_inferred': True}
    return None

def _year_grouped_periods(blocks, months, as_of):
    """Assign (year, month, kind) to each month column. Year walks forward across
    a block (JAN rollover -> +1) starting from the block's base year; kind from
    the block label (ACT -> actual, FORECAST/PLAN -> forecast)."""
    blocks = sorted(blocks)
    starts = [(c, yb) for c, yb in blocks]
    def block_for(col):
        chosen = None
        for c, yb in starts:
            if c <= col:
                chosen = yb
            else:
                break
        return chosen
    periods = {}
    # group month columns by their owning block to reset the year walk per block
    by_block = defaultdict(list)
    for ci, mn, ey in months:
        b = None
        for c, yb in starts:
            if c <= ci:
                b = c
            else:
                break
        by_block[b].append((ci, mn, ey))
    # Process blocks IN COLUMN ORDER so a yearless block can thread its base year from
    # the previous block's end (Dec -> Jan rollover), instead of defaulting to as_of.year.
    # A header can label the actual year(s) explicitly on some blocks ('ACT 2024',
    # 'Plan 2026') but leave others bare ('Actual'); a bare block sitting right after a
    # block that ends in December and itself starting in January is that next year.
    running_end = None   # (year, month) assigned to the previous block's last column
    for bstart in sorted(by_block.keys(), key=lambda x: (x is None, x)):
        cols = by_block[bstart]
        yb = dict(starts).get(bstart)
        if yb is None:
            yb = {'kind': 'actual', 'years': []}
        first_mn = sorted(cols)[0][1]
        if yb['years']:
            base = yb['years'][0]
        elif running_end is not None and running_end[1] == 12 and first_mn == 1:
            base = running_end[0] + 1          # clean Dec -> Jan rollover: prior year + 1
        else:
            base = as_of.year
        # provisional forward-walk (JAN rollover -> +1)
        prov = {}
        prev = 0
        y = base
        for ci, mn, _ey in sorted(cols):
            if prev and mn < prev:
                y += 1
            prev = mn
            prov[ci] = y
        # snap the block to any EXPLICIT year token (e.g. JAN'26) so a block that
        # starts in a prior year then crosses into the cutover year lands right.
        delta = 0
        for ci, mn, ey in sorted(cols):
            if ey is not None:
                delta = ey - prov[ci]
                break
        for ci, mn, _ey in cols:
            periods[ci] = (prov[ci] + delta, mn, yb['kind'])
        last_ci, last_mn, _ = sorted(cols)[-1]
        running_end = (prov[last_ci] + delta, last_mn)
    return periods

# ---- line resolution (section + direction state machine) --------------------
SECTION_MAP = [
    ('NONOPERATIONAL', 'Non Operational'),
    ('NONOPERATION', 'Non Operational'),
    ('OPERATION', 'Operation'),
    ('OPERATIONS', 'Operation'),
    ('INTEREST', 'Interest'),
    ('WITHINGROUP', 'Within Group'),
    ('TRANSFERS', 'Within Group'),
    ('TRANSFER', 'Within Group'),
    ('BANKFINANCING', 'Bank Financing'),
    ('BANKING', 'Bank Financing'),
    ('NEWSALES', 'New Sales'),
]
# Exact label-column banner phrases that start a section (matched regardless of any
# subtotal the banner row carries — unlike the section column, these banners can sit
# in the label column with no section tag, e.g. KSA 'WITHIN GROUP' / 'BANK FINANCING').
SECTION_BANNERS = {
    'OPERATION': 'Operation', 'OPERATIONS': 'Operation',
    'NONOPERATIONAL': 'Non Operational', 'NONOPERATION': 'Non Operational',
    'INTEREST': 'Interest', 'WITHINGROUP': 'Within Group',
    'BANKFINANCING': 'Bank Financing', 'BANKING': 'Bank Financing',
    'NEWSALES': 'New Sales',
}
SKIP_LABELS = {'TOTAL', 'NET', 'NETFUNDS', 'NETBANKINGFINANCE', 'NETBALANCEATSTART',
               'NETTRANSFERS', 'NETTRANSFERSWITHINAREA', 'NETTRANSFEROUTSIDEAREA',
               'NETTRANSFERSWITHINAREAOUTSIDE', 'NETTRANSFERMOA', 'DESCRIPTION',
               'NETCASHINOUT', 'NETCASH', 'RECEIPTS', 'PAYMENTS',
               'NETTRANSFEROUTSIDEAREAMOA', 'NETTRANSFERMOA', 'NETTRANSFERSMOA',
               'TOTALCASHCOLLATERAL', 'TOTALCASHCOLATERAL', 'CASH', 'GRANDTOTAL',
               'PROJECT', 'AREADESCRIPTION'}
OPENING = {'OPENINGBALANCE', 'CASHOPENINGBALANCE', 'BALANCEATSTART',
           'BALANCEATSTARTCASH', 'BALANCEATSTARTCASH000'}
ENDING = {'ENDINGBALANCE', 'CASHENDBALANCE', 'BALANCEATEND', 'CASHENDBALANCE000',
          'BALANCEATENDUSD000', 'BALANCEATENDLIQUIDFUNDS', 'BALANCEATENDLIQUIDFUND'}


def is_opening_label(lt):
    """True if a tightened label names an opening/start-of-period cash balance.
    Excludes 'NET BALANCE AT START' (a memo subtotal, in SKIP_LABELS)."""
    if lt in SKIP_LABELS:
        return False
    return lt in OPENING or lt.startswith('BALANCEATSTART') \
        or ('OPENING' in lt and 'BALANCE' in lt)


def is_ending_label(lt):
    """True if a tightened label names an ending/closing cash balance. Note the
    rollup often carries TWO ending rows (native liquid-funds then a USD '000
    restatement); callers that want a single series take the FIRST match."""
    if lt in SKIP_LABELS:
        return False
    return lt in ENDING or lt.startswith('BALANCEATEND') \
        or (('ENDING' in lt or 'CLOSING' in lt) and 'BALANCE' in lt)


# Sentinel for an INTENTIONAL skip (a subtotal/memo row we deliberately don't map),
# as distinct from None (a genuine miss worth flagging in "Lines that didn't map").
# resolve_line returns this for rows like the Bank Financing 'TOTAL …' subtotals; the
# parse loop drops them silently instead of listing them as unmatched.
SKIP = object()


def section_of(label_tight):
    for key, cat in SECTION_MAP:
        if label_tight.startswith(key) or label_tight == key:
            return cat
    return None

def resolve_line(label, section, direction, post_ending, resolver, area, wg_sub=None):
    """Return (line_code or None, new_direction). `direction` is 'Receipts'/'Payments'.
    wg_sub carries the active Within-Group subsection ('outside_area'/'within_area'/
    'treasury'/'moa') set by a banner row, so counterparty-named detail rows under it
    (e.g. CCUW's 'UAE - CCEPTE', 'ADCB Legal Fees' under 'PAYMENTS - OUTSIDE AREA')
    inherit the classification instead of falling through to unmatched."""
    lt = tight(label)
    # In Interest / Non-Operational a bare RECEIPTS / PAYMENTS row IS the data line,
    # so don't let the SKIP_LABELS guard reject it (it stays a skippable sub-header
    # everywhere else — Operation etc.).
    section_data = (section in ('Interest', 'Non Operational', 'Within Group')
                    and lt in ('RECEIPTS', 'RECEIPT', 'PAYMENTS', 'PAYMENT'))
    if not lt or (lt in SKIP_LABELS and not section_data):
        return None, direction
    # balance lines (exact set, then label-variant keyword fallback, e.g.
    # "Opening bank balance ($'mil)" / "Ending bank balance")
    if lt in OPENING:
        return 'opening_balance', direction
    if lt in ENDING:
        return 'ending_balance', direction
    if 'OPENING' in lt and 'BALANCE' in lt:
        return 'opening_balance', direction
    if ('ENDING' in lt or 'CLOSING' in lt) and 'BALANCE' in lt:
        return 'ending_balance', direction
    if post_ending:
        if lt.startswith('LOAN') or lt == 'TREASURYLOANS' \
                or 'ACCUMULATEDLOAN' in lt or 'ACCOUMULATEDLOAN' in lt:
            return 'accum_loans', direction
        if lt.startswith('OVERDRAFT') \
                or 'ACCUMULATEDOVERDRAFT' in lt or 'ACCOUMULATEDOVERDRAFT' in lt:
            return 'accum_od', direction
        # Nothing below the ending balance is a FLOW line — it's accumulated stocks
        # (handled above) or memo/notes. Don't fall through to the section blocks, or
        # a note that happens to carry a stray value and a flow keyword gets
        # mis-mapped (e.g. KAZH's '- $ 200K Loan from KCC in May'25' note, with a
        # leftover value, resolving to bf_pay_loans and inflating the file-side bank
        # section in the compare-to-file view). Return None for everything else.
        return None, direction

    # --- Within Group (transfers) ---
    if section == 'Within Group':
        d = direction
        if lt.startswith('RECEIPT'):
            d = 'Receipts'
        elif lt.startswith('PAYMENT'):
            d = 'Payments'
        nat = d or 'Receipts'
        if 'MOA' in lt:
            code = 'wg_recpt_moa' if nat == 'Receipts' else 'wg_pay_moa'
        elif 'OUTSIDE' in lt or 'OTHERAREA' in lt or 'OTHERAREAS' in lt:
            code = 'wg_recpt_outside_area' if nat == 'Receipts' else 'wg_pay_outside_area'
        elif 'TREASURY' in lt:
            code = 'wg_recpt_treasury' if nat == 'Receipts' else 'wg_pay_treasury'
        elif 'FROMAREAS' in lt:
            code = 'treasury_recpt_areas'
        elif 'TOAREAS' in lt:
            code = 'treasury_pay_areas'
        elif 'WITHIN' in lt or 'AREA' in lt or tight(area) in lt or lt in ('QATAR',):
            code = 'wg_recpt_within_area' if nat == 'Receipts' else 'wg_pay_within_area'
        elif wg_sub:
            # detail row with no keyword of its own — inherit the banner's subsection
            sub = {'outside_area': ('wg_recpt_outside_area', 'wg_pay_outside_area'),
                   'within_area':  ('wg_recpt_within_area',  'wg_pay_within_area'),
                   'treasury':     ('wg_recpt_treasury',     'wg_pay_treasury'),
                   'moa':          ('wg_recpt_moa',          'wg_pay_moa')}[wg_sub]
            code = sub[0] if nat == 'Receipts' else sub[1]
        else:
            return None, d
        return code, d

    # --- Bank Financing ---
    if section == 'Bank Financing':
        # 'TOTAL LOANS IN' / 'TOTAL OVERDRAFT' / 'TOTAL LOANS OUT' are subtotals of
        # the detail rows (LOAN IN/OUT, Overdraft In/Out) — skip so the section sums
        # to the true NET banking finance, not double. Returned as SKIP (not None) so
        # the parse loop drops them silently rather than flagging them as unmatched.
        if lt.startswith('TOTAL'):
            return SKIP, direction
        d = direction
        # direction can lead or trail the label: 'IN LOANS' / 'TOTAL LOANS IN',
        # 'OUT LOAN' / 'LOAN OUT' / 'Overdraft Out'.
        if lt.startswith('OUT') or lt.endswith('OUT'):
            d = 'Payments'
        elif lt.startswith('IN') or lt.endswith('IN'):
            d = 'Receipts'
        elif lt.startswith('RECEIPT'):
            d = 'Receipts'
        elif lt.startswith('PAY'):
            d = 'Payments'
        if 'SETTLED' in lt:
            d = 'Payments'   # repayment / unwind
        nat = d or 'Receipts'
        if 'LTR' in lt or 'DISCOUNTED' in lt or 'TRUST' in lt:
            code = 'bf_recpt_discounted' if nat == 'Receipts' else 'bf_pay_discounted'
        elif 'OVERDRAFT' in lt:
            code = 'bf_recpt_od' if nat == 'Receipts' else 'bf_pay_od'
        elif 'LOAN' in lt:
            code = 'bf_recpt_loans' if nat == 'Receipts' else 'bf_pay_loans'
        else:
            return None, d
        return code, d

    # --- Interest / Non-Operational: RECEIPT / PAYMENT are the data lines ---
    if section in ('Interest', 'Non Operational'):
        if lt.startswith('RECEIPT'):
            code = 'interest_recpt' if section == 'Interest' else 'nonop_recpt'
            return code, 'Receipts'
        if lt.startswith('PAYMENT'):
            code = 'interest_pay' if section == 'Interest' else 'nonop_pay'
            return code, 'Payments'

    # --- Interest embedded in Operation -> its own Interest section ---
    # Some area files list INTEREST as an operation receipt/payment line rather than
    # under a dedicated INTEREST banner (e.g. Kazakhstan). Pull it into the canonical
    # Interest section by label + direction, symmetrically on the rollup and the
    # project sheets, so the section total is consistent and the ending balance still
    # ties. Scoped to Operation/None so New Sales interest and the dedicated Interest
    # section (handled above) are left alone.
    if section in ('Operation', None) and lt.startswith('INTEREST'):
        return ('interest_recpt' if (direction or 'Receipts') == 'Receipts'
                else 'interest_pay'), direction

    # --- dictionary lookup for Operation / New Sales / misc ---
    nat = direction or 'Receipts'
    if lt == 'CAPEX':
        return 'oper_pay_capex', 'Payments'
    if lt == 'CIT':
        return 'oper_pay_cit', 'Payments'
    code = resolver.lookup(label, nat, section or 'Operation')
    if code:
        return code, direction
    # retry with parenthetical qualifiers stripped, e.g.
    # 'SUPPLIERS - LOCAL (include blasting)' -> 'SUPPLIERS - LOCAL'
    if '(' in label:
        base = re.sub(r'\(.*?\)', '', label).strip()
        if base and tight(base) != lt:
            code = resolver.lookup(base, nat, section or 'Operation')
            if code:
                return code, direction

    # --- keyword fallback (year-grouped sheets where the section column is not
    #     reliably tagged: resolve banking / JV / claims by label keyword with
    #     embedded direction). Fires only after section + dictionary both miss. ---
    code, d = keyword_fallback(lt, direction)
    if code:
        return code, d
    return None, direction

def keyword_fallback(lt, direction):
    """Resolve the cross-area banking/JV/claims cluster from label keywords.
    Direction is embedded in the label (FROM/IN vs TO/OUT/PAY/SETTLED)."""
    is_out = any(k in lt for k in ('TO', 'OUT', 'PAY', 'SETTLE'))
    is_in = any(k in lt for k in ('FROM', 'IN', 'RECEIPT'))
    # within-group transfers (year-grouped sheets where the section column isn't
    # tagged TRANSFERS): 'RECEIPTS - WITHIN AREA', 'PAYMENTS - OUTSIDE AREA MOA', ...
    if ('WITHINAREA' in lt or 'OUTSIDEAREA' in lt) and \
            (lt.startswith('RECEIPT') or lt.startswith('PAYMENT')):
        nat = 'Payments' if lt.startswith('PAYMENT') else 'Receipts'
        if 'MOA' in lt:
            return ('wg_recpt_moa' if nat == 'Receipts' else 'wg_pay_moa'), nat
        if 'OUTSIDEAREA' in lt:
            return ('wg_recpt_outside_area' if nat == 'Receipts' else 'wg_pay_outside_area'), nat
        return ('wg_recpt_within_area' if nat == 'Receipts' else 'wg_pay_within_area'), nat
    # accumulated balances (closing block)
    if 'ACCOUMULATEDLOAN' in lt or 'ACCUMULATEDLOAN' in lt:
        return 'accum_loans', direction
    if lt in ('TOTALOVERDRAFT', 'ACCUMULATEDOVERDRAFT', 'ACCOUMULATEDOVERDRAFT'):
        return 'accum_od', direction
    # JV
    if 'JV' in lt:
        if lt.startswith('TO') or is_out and not is_in:
            return 'oper_pay_jv', 'Payments'
        return 'oper_recpt_jv', 'Receipts'
    # claims (approved / expected / unfinalised)
    if 'CLAIM' in lt:
        return 'claims_recpt', 'Receipts'
    # discounted invoices / LTR / trust receipts
    if 'DISCOUNTED' in lt or 'LTR' in lt or 'TRUST' in lt:
        return ('bf_pay_discounted', 'Payments') if (lt.startswith('PAY') or is_out) \
            else ('bf_recpt_discounted', 'Receipts')
    # overdrafts (banking, not accumulated)
    if 'OVERDRAFT' in lt:
        if 'SETTLE' in lt or lt.startswith('OUT') or 'OVERDRAFTOUT' in lt:
            return 'bf_pay_od', 'Payments'
        return 'bf_recpt_od', 'Receipts'
    # loans (banking)
    if 'LOAN' in lt:
        if 'SETTLE' in lt or lt.startswith('OUT') or 'LOANOUT' in lt or 'LOANSOUT' in lt or lt == 'TOTALLOANSOUT':
            return 'bf_pay_loans', 'Payments'
        return 'bf_recpt_loans', 'Receipts'
    return None, direction

# Areas whose files append a REDUNDANT duplicate period block that maps to periods
# already covered by an earlier column (EPSO repeats JAN-APR 2026 under a second
# 'FORECAST 2026' header after JAN'26..DEC'26). For these, keep only the first column
# per (year, month) so the duplicate doesn't sum (doubling opening/flows). Scoped by
# area — most files legitimately carry two columns per period (native + USD, actual +
# forecast) that must NOT be deduped, so this is opt-in per area.
DEDUP_PERIOD_AREAS = {'EPSO'}


def _dedup_periods(periods):
    seen, out = set(), {}
    for ci in sorted(periods):
        ym = (periods[ci][0], periods[ci][1])
        if ym in seen:
            continue
        seen.add(ym)
        out[ci] = periods[ci]
    return out


# ---- per-sheet parse --------------------------------------------------------
def parse_sheet(ws, area, sheet_name, resolver, as_of, verbose=False, wg_sign_flip=True):
    rows = list(ws.iter_rows(min_row=1, max_row=80, max_col=120, values_only=True))
    hdr = detect_header(rows, as_of)
    if not hdr:
        return None, {'sheet': sheet_name, 'reason': 'no period axis'}
    lc, sc = hdr['label_col'], hdr['section_col']
    periods = hdr['periods']
    if area in DEDUP_PERIOD_AREAS:
        periods = _dedup_periods(periods)
    project_code, _cur = strip_currency(sheet_name)
    project_code = project_code or sheet_name.strip()

    out_rows = []
    unmatched = {}   # label -> {count, cells: ["Sheet!C12", ...]} for reviewer lookup
    direction = None
    section = None
    wg_sub = None    # active Within-Group subsection (set by a banner, applied to detail)
    post_ending = False
    start = hdr['header_row'] + 1

    for i, r in enumerate(rows[start:], start=start):
        label = r[lc] if lc < len(r) and isinstance(r[lc], str) else None
        seccell = r[sc] if sc < len(r) and isinstance(r[sc], str) else None
        # update section from the section column
        if seccell:
            cat = section_of(tight(seccell))
            if cat:
                section = cat
                direction = None
                wg_sub = None
        if not label or not label.strip():
            continue
        lt = tight(label)
        # A section banner can sit in the LABEL column (no section-column tag) —
        # e.g. 'WITHIN GROUP' / 'BANK FINANCING'. An EXACT banner phrase flips the
        # section even if the banner row also carries a subtotal (its value is
        # skipped); a longer header phrase flips only when it carries no period
        # values, so a data row like 'OPERATION & MAINTENANCE' is never mistaken
        # for a header. Without this the section stays on the prior one and the
        # whole block (within-group / banking) is swallowed or dropped.
        lab_sec = SECTION_BANNERS.get(lt)
        if not lab_sec:
            cand = section_of(lt)
            if cand and not any(isinstance(r[ci], (int, float)) and r[ci] not in (0, None)
                                for ci in periods if ci < len(r)):
                lab_sec = cand
        # 'INTEREST' doubles as a section header AND an Operation line item — some
        # files list an INTEREST receipt line and an INTEREST payment line INSIDE the
        # Operation block (e.g. Kazakhstan), rather than a dedicated Interest section.
        # When we're already inside an Operation / New Sales flow (a direction is set),
        # treat INTEREST as a data line, NOT a header: don't let it flip the section
        # (which would mis-bucket every following line and drop the interest itself).
        # resolve_line then maps it to interest_recpt / interest_pay by direction.
        # A genuine dedicated Interest section is marked in the section column, so it
        # still flips correctly. Applies to the empty receipt line too, so it doesn't
        # hijack the section before the payment line is reached.
        if lab_sec == 'Interest' and lt == 'INTEREST' \
                and section in ('Operation', 'New Sales') and direction:
            lab_sec = None
        if lab_sec:
            section = lab_sec
            direction = None
            wg_sub = None
            continue
        # pure direction sub-headers
        if lt in ('RECEIPTS', 'RECEIPT') and section in ('Operation', 'New Sales', None):
            direction = 'Receipts'; continue
        if lt in ('PAYMENTS', 'PAYMENT') and section in ('Operation', 'New Sales', None):
            direction = 'Payments'; continue
        # bare banking direction sub-headers
        if lt == 'IN':
            direction = 'Receipts'; continue
        if lt in ('OUT', 'OUTSETTLE'):
            direction = 'Payments'; continue
        # Within-Group subsection banner (e.g. 'RECEIPTS - OUTSIDE AREA',
        # 'PAYMENTS - OUTSIDE AREA'): a NO-VALUE row naming a transfer subsection
        # (OUTSIDE/WITHIN AREA, TREASURY, MOA) + optional direction. Records the
        # subsection so counterparty detail rows beneath (e.g. CCUW 'UAE - CCEPTE',
        # 'ADCB Legal Fees') inherit the classification, then skips the banner. Also
        # ESTABLISHES the Within Group section — needed for sheets whose section title
        # sits above the detected header (e.g. CCUW's 'TRANSFERS OUTSIDE AREA' sheet).
        # No-value-only so a valued line still maps by its own keyword (no double-count).
        wsub = ('moa' if 'MOA' in lt else
                'outside_area' if 'OUTSIDEAREA' in lt else
                'within_area' if 'WITHINAREA' in lt else
                'treasury' if 'TREASURY' in lt else None)
        if wsub:
            row_has_val = any(isinstance(r[ci], (int, float)) and r[ci] not in (0, None)
                              for ci in periods if ci < len(r))
            if not row_has_val:
                section = 'Within Group'
                wg_sub = wsub
                if lt.startswith('RECEIPT'):
                    direction = 'Receipts'
                elif lt.startswith('PAYMENT'):
                    direction = 'Payments'
                continue
        # In Interest / Non-Operational the bare RECEIPTS / PAYMENTS rows ARE the
        # data lines (one receipt + one payment carry the section's whole value) —
        # unlike Operation, where they are sub-headers. So don't let SKIP_LABELS
        # drop them; fall through to resolve_line, which maps them to
        # interest_recpt/pay and nonop_recpt/pay. Within Group is the same: a bare
        # 'PAYMENTS -' row under a subsection banner (e.g. CCUW's 'RECEIPTS - WITHIN
        # AREA' → 'PAYMENTS -') is a real transfer line that resolves via wg_sub —
        # without this it tightens to 'PAYMENTS' ∈ SKIP_LABELS and the within-area
        # payments are silently dropped (the OUTSIDE/TREASURY rows survive because
        # their label keeps the keyword).
        is_section_data = (section in ('Interest', 'Non Operational', 'Within Group')
                           and lt in ('RECEIPTS', 'RECEIPT', 'PAYMENTS', 'PAYMENT'))
        if lt in SKIP_LABELS and not is_section_data:
            continue
        code, direction = resolve_line(label, section, direction, post_ending,
                                       resolver, area, wg_sub)
        if lt in ENDING or (('ENDING' in lt or 'CLOSING' in lt) and 'BALANCE' in lt):
            post_ending = True
        # Intentional skip (e.g. a Bank Financing 'TOTAL …' subtotal) — drop it
        # silently; it's not a genuine miss, so it doesn't belong in the unmatched list.
        if code is SKIP:
            continue
        if code is None:
            # only report labels that actually carry data (ignore stray text)
            has_val = any(isinstance(r[ci], (int, float)) and r[ci] not in (0, None)
                          for ci in periods if ci < len(r))
            if has_val or lt not in ('', 'P R O J E C T'):
                # rows[i] is spreadsheet row i+1; lc is 0-based -> column letter
                cell = f"{sheet_name}!{get_column_letter(lc + 1)}{i + 1}"
                ent = unmatched.setdefault(label.strip(), {'count': 0, 'cells': []})
                ent['count'] += 1
                if cell not in ent['cells'] and len(ent['cells']) < 5:
                    ent['cells'].append(cell)
            continue
        for ci, (yy, mm, kind) in periods.items():
            if yy < MIN_YEAR or ci >= len(r):
                continue
            v = r[ci]
            if not isinstance(v, (int, float)) or v == 0:
                continue
            out_rows.append({'area': area, 'project_code': project_code,
                             'line_code': code, 'year': yy, 'month': mm,
                             'value': round(float(v), 4), 'kind': kind})

    # Sign normalisation — outflows must be stored NEGATIVE so a section nets by
    # summing. Area files are inconsistent: some list operation/interest/non-op
    # PAYMENTS as positive magnitudes (e.g. Kazakhstan), others store them already
    # signed (e.g. Saudi). Within Group / Bank Financing are signed in every file
    # (their detail rows carry the minus), so they're excluded. Detect per sheet
    # from the signed sum of the flippable payment lines: if it's positive the file
    # used gross magnitudes, so negate them — multiplying by -1 (not -abs) keeps a
    # genuine negative refund flipping to a positive inflow. Applied identically to
    # project sheets and the rollup, so Σ-projects vs rollup still reconciles.
    flip_cats = {'Operation', 'New Sales', 'Interest', 'Non Operational'}

    def _flip_pay(c):
        return (resolver.nat_by_code.get(c) == 'Payments'
                and resolver.cat_by_code.get(c) in flip_cats)
    pay_sum = sum(rr['value'] for rr in out_rows if _flip_pay(rr['line_code']))
    if pay_sum > 0:
        for rr in out_rows:
            if _flip_pay(rr['line_code']):
                rr['value'] = round(-rr['value'], 4)

    # Within Group gets its OWN per-sheet sign detection (NOT folded into the flip
    # above): most files sign their transfer detail (KSA/KZ carry the minus), but some
    # store the transfer-OUT detail as gross positive under a PAYMENTS banner (e.g.
    # CCUW's 'TRANSFERS OUTSIDE AREA' sheet). A single combined detection would break
    # Kazakhstan, where one sheet mixes GROSS operations with already-SIGNED
    # within-group. Detect on the within-group payment lines alone, per sheet.
    def _wg_pay(c):
        return (resolver.nat_by_code.get(c) == 'Payments'
                and resolver.cat_by_code.get(c) == 'Within Group')
    # A sheet can carry a legitimate SIGNED-positive within-group payment (a transfer
    # reversal) amid an otherwise signed convention — indistinguishable from a gross
    # magnitude by the sheet alone. When the area's rollup proves the whole area is
    # signed (e.g. Qatar: CONSOL within-area pay is negative and nets to zero only if
    # the lone positive stays positive), the caller disables the flip via wg_sign_flip.
    wg_pay_sum = sum(rr['value'] for rr in out_rows if _wg_pay(rr['line_code']))
    if wg_sign_flip and wg_pay_sum > 0:
        for rr in out_rows:
            if _wg_pay(rr['line_code']):
                rr['value'] = round(-rr['value'], 4)

    meta = {'sheet': sheet_name, 'project_code': project_code, 'mode': hdr['mode'],
            'label_col': lc, 'n_rows': len(out_rows), 'unmatched': dict(unmatched),
            'pay_flipped': pay_sum > 0}
    return out_rows, meta

# ---- per-workbook parse -----------------------------------------------------
def pick_usd_sheets(wb):
    """Return list of (sheet_name) that are USD project sheets.
    Prefer explicit ' USD' suffix; if a file has no USD-suffixed project sheet
    at all, fall back to all non-rollup sheets (single-currency-per-sheet file)."""
    names = wb.sheetnames
    has_usd_suffix = any(strip_currency(n)[1] == 'USD' for n in names)
    sel = []
    for n in names:
        cls, cur = classify(n)
        if cls != 'project':
            continue
        if has_usd_suffix:
            if cur == 'USD':
                sel.append(n)
        else:
            # single-currency file: take project sheets that aren't a local dup
            if cur in (None, 'USD'):
                sel.append(n)
    return sel

def parse_workbook(path, resolver, as_of, verbose=False):
    fn = os.path.basename(path)
    area = AREA_BY_FILE.get(fn, fn.split('_')[0])
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sel = pick_usd_sheets(wb)
    all_rows, sheet_meta, no_axis = [], [], []
    unmatched_labels = Counter()
    projects = {}
    for n in sel:
        rows, meta = parse_sheet(wb[n], area, n, resolver, as_of, verbose)
        if rows is None:
            no_axis.append(meta); continue
        all_rows.extend(rows)
        sheet_meta.append(meta)
        for lab, c in meta['unmatched'].items():
            unmatched_labels[lab] += c
        pc = meta['project_code']
        gi = resolver.gacc.get(tight(pc))
        projects[pc] = {'sheet': n, 'rows': meta['n_rows'],
                        'gacc': gi['abbreviation'] if gi else None,
                        'via': gi['via'] if gi else None}
    wb.close()
    unmatched_projects = {pc: m for pc, m in projects.items() if not m['gacc']}
    return {
        'area': area, 'file': fn, 'n_usd_sheets': len(sel),
        'n_no_axis': len(no_axis), 'no_axis': no_axis,
        'rows': all_rows, 'sheet_meta': sheet_meta,
        'projects': projects, 'unmatched_projects': unmatched_projects,
        'unmatched_labels': dict(unmatched_labels.most_common()),
    }

# ---- CLI --------------------------------------------------------------------
def fmt_summary(res):
    L = []
    L.append(f"=== {res['area']} ({res['file']}) ===")
    L.append(f"USD project sheets: {res['n_usd_sheets']} | parsed rows: {len(res['rows'])} "
             f"| sheets w/o axis: {res['n_no_axis']}")
    nproj = len(res['projects']); nmatch = nproj - len(res['unmatched_projects'])
    L.append(f"projects: {nproj} ({nmatch} linked to gacc, {len(res['unmatched_projects'])} unmatched)")
    ka = sum(1 for r in res['rows'] if r['kind'] == 'actual')
    kf = len(res['rows']) - ka
    yrs = sorted({r['year'] for r in res['rows']})
    L.append(f"actual rows: {ka} | forecast rows: {kf} | years: {yrs}")
    if res['unmatched_labels']:
        L.append(f"UNMATCHED LABELS ({len(res['unmatched_labels'])}): " +
                 ", ".join(f"{k!r}x{v}" for k, v in list(res['unmatched_labels'].items())[:25]))
    else:
        L.append("UNMATCHED LABELS: none")
    if res['unmatched_projects']:
        L.append(f"UNMATCHED PROJECTS ({len(res['unmatched_projects'])}): " +
                 ", ".join(sorted(res['unmatched_projects'])))
    if res['no_axis']:
        L.append("NO-AXIS SHEETS: " + ", ".join(m['sheet'] for m in res['no_axis']))
    return "\n".join(L)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('file', nargs='?', help='workbook filename (in SRC) or path')
    ap.add_argument('--all', action='store_true', help='sweep every workbook')
    ap.add_argument('--as-of', default=None, help='actual/forecast cutover YYYY-MM-DD')
    ap.add_argument('--out', default=None, help='write normalized rows + reports to JSON')
    ap.add_argument('-v', '--verbose', action='store_true')
    args = ap.parse_args()

    as_of = (datetime.date.fromisoformat(args.as_of) if args.as_of else DEFAULT_AS_OF)
    resolver = Resolver(load_ref())

    if args.all:
        files = sorted(glob.glob(SRC + '*.xlsx'))
        grand = {'as_of': str(as_of), 'areas': []}
        for p in files:
            res = parse_workbook(p, resolver, as_of, args.verbose)
            print(fmt_summary(res)); print()
            grand['areas'].append({k: res[k] for k in
                ('area','file','n_usd_sheets','n_no_axis','unmatched_labels','unmatched_projects')})
            grand['areas'][-1]['n_rows'] = len(res['rows'])
        if args.out:
            with open(args.out, 'w') as f:
                json.dump(grand, f, indent=1)
            print("Wrote", args.out)
        return

    if not args.file:
        ap.error("give a workbook filename or --all")
    path = args.file if os.path.exists(args.file) else SRC + args.file
    res = parse_workbook(path, resolver, as_of, args.verbose)
    print(fmt_summary(res))
    if args.out:
        with open(args.out, 'w') as f:
            json.dump({'as_of': str(as_of), 'area': res['area'], 'file': res['file'],
                       'rows': res['rows'], 'sheet_meta': res['sheet_meta'],
                       'projects': res['projects'],
                       'unmatched_labels': res['unmatched_labels'],
                       'unmatched_projects': res['unmatched_projects']}, f, indent=1)
        print("\nWrote", args.out, f"({len(res['rows'])} rows)")

if __name__ == '__main__':
    main()
