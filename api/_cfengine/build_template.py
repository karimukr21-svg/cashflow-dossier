#!/usr/bin/env python3
"""Cash-flow SUBMISSION TEMPLATE generator — the input side of the pipeline.

Generates the locked, pre-filled Excel workbook that Treasury / areas / projects fill in
each cycle. One workbook per area: a sheet per project (exactly the project codes the
canonical data already carries) plus a formula-only SUMMARY rollup.

THE DESIGN RULE — the template is DERIVED, never authored.
    Row skeleton      <- cf_lines (category / nature / sort_order), the canonical taxonomy
    Sheet list        <- distinct project_code for the area in the cycle version
    Pre-filled actual <- cf_actuals   (periods <= as-of)   LOCKED, grey
    Pre-filled fcst   <- cf_forecasts (periods >  as-of)   OPEN, white  <- their input
There is no per-area structure config here on purpose. The parser configs in
reconcile_stage.py describe how to READ messy files; this describes how files should BE.
Deriving from published data means the file that comes back next cycle matches the DB
because the DB generated it.

WHAT THE LOCKING BUYS (every real defect of the June cycle, structurally eliminated):
  * SUMMARY is 100% formulas over the project sheets and is protected -> nobody can type
    a constant over a rollup formula (3 of 4 real June defects were exactly that).
  * The AREA sheet is the SANCTIONED place for area-level entries. The hardcodes happened
    because someone had a real number and no legitimate row for it. Now there is one.
  * Opening/Ending balances are chained by formula (ending = opening + net movement,
    next opening = prior ending) -> stale per-project balance rows cannot drift.
  * A hidden _meta sheet carries area / cycle / version / scale / per-sheet project code,
    so intake is label-anchored and mechanical rather than cell-position guesswork.

Sheet protection is deterrence + guidance, not security (Excel protection is trivially
removed). The point is that the correct path is the easy one.
"""
from __future__ import annotations

import datetime as dt
import io
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

import db

TEMPLATE_VERSION = "1.0"

# ---- CCC brand -------------------------------------------------------------------
# NOTE: cell text is Calibri on purpose. Rubik is the CCC brand face but it is not
# installed on coordinator machines, so Excel would silently substitute and the file
# would look broken on their screen. Brand identity is carried by COLOUR, not the face.
CRIMSON = "E10020"
CHARCOAL = "141414"
WHITE = "FFFFFF"
GREY_BAND = "F4F4F4"      # locked actuals
GREY_RULE = "D9D9D9"
INPUT_TINT = "FFF3C4"     # their input cells — soft yellow, the "fill me" convention
INPUT_RULE = "E9CE6A"     # a warmer border on input cells so the region reads at a glance
TOTAL_TINT = "EFEFEF"
FONT = "Calibri"

THIN = Side(style="thin", color=GREY_RULE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
_INPUT_SIDE = Side(style="thin", color=INPUT_RULE)
INPUT_BOX = Border(left=_INPUT_SIDE, right=_INPUT_SIDE, top=_INPUT_SIDE, bottom=_INPUT_SIDE)

NUM_FMT = "#,##0;[Red](#,##0)"

# Balance lines are STOCKS, not flows - they must never be summed down the statement.
BALANCE_NATURE = "Balance"
OPENING = "opening_balance"
ENDING = "ending_balance"
MEMO_STOCKS = ("accum_loans", "accum_od")


# =================================================================================
# Data
# =================================================================================
def load_lines():
    """Canonical row taxonomy, already ordered. Inactive (_lgc legacy) lines dropped."""
    rows = db.select("cf_lines", {
        "select": "line_code,nature,category,description,sign_convention,sort_order",
        "is_active": "eq.true",
        "order": "sort_order.asc",
    })
    return rows


def load_cells(area, version, cycle_year, as_of_ym):
    """Pre-fill values keyed (project_code, line_code, year, month).

    actuals  = published truth for elapsed periods -> locked in the sheet
    forecast = the selected cycle version for future periods -> their input
    """
    fc = db.select("cf_forecasts", {
        "select": "project_code,line_code,year,month,value,currency",
        "version": f"eq.{version}",
        "area": f"eq.{area}",
    })
    ac = db.select("cf_actuals", {
        "select": "project_code,line_code,year,month,value,currency",
        "area": f"eq.{area}",
        "year": f"gte.{cycle_year}",
    })

    actual, forecast, ccy = {}, {}, None
    for r in ac:
        ym = r["year"] * 100 + r["month"]
        if ym > as_of_ym:
            continue
        actual[(r["project_code"], r["line_code"], r["year"], r["month"])] = r["value"]
        ccy = ccy or r.get("currency")
    for r in fc:
        ym = r["year"] * 100 + r["month"]
        ccy = ccy or r.get("currency")
        if ym <= as_of_ym:
            # elapsed: actuals win. Keep as fallback only if nothing was published.
            key = (r["project_code"], r["line_code"], r["year"], r["month"])
            actual.setdefault(key, r["value"])
            continue
        forecast[(r["project_code"], r["line_code"], r["year"], r["month"])] = r["value"]

    years = sorted({r["year"] for r in fc if r["year"] >= cycle_year})
    projects = sorted({r["project_code"] for r in fc} | {r["project_code"] for r in ac})
    return actual, forecast, ccy or "USD", years or [cycle_year], projects


# =================================================================================
# Layout
# =================================================================================
def _col_plan(years):
    """Column layout: 12 months then a Total for each year.

    One grand total across a 3-year sheet is meaningless for flows and isn't how the
    coordinator files read — each year closes with its own total.
    Returns entries ('m', year, month) | ('t', year, None); index i -> column 2 + i.
    """
    plan = []
    for y in years:
        for m in range(1, 13):
            plan.append(("m", y, m))
        plan.append(("t", y, None))
    return plan


def _row_plan(lines):
    """Ordered render plan: category bands, Receipts/Payments sub-blocks, net rows.

    Returns a list of ('band'|'sub'|'line'|'net', payload) in statement order, covering
    FLOW lines only. Balances are placed by the writer (opening at top, ending at the
    bottom, stock memos below) because they are stocks and behave differently.
    """
    flows = [l for l in lines if l["nature"] != BALANCE_NATURE]
    by_cat = defaultdict(list)
    for l in flows:
        by_cat[l["category"]].append(l)

    cats = sorted(by_cat, key=lambda c: min(l["sort_order"] for l in by_cat[c]))
    plan = []
    for cat in cats:
        plan.append(("band", cat))
        for nature in ("Receipts", "Payments"):
            block = [l for l in by_cat[cat] if l["nature"] == nature]
            if not block:
                continue
            plan.append(("sub", nature))
            for l in sorted(block, key=lambda x: x["sort_order"]):
                plan.append(("line", l))
        plan.append(("net", cat))
    return plan


def _style_header(ws, ncols, area, project, ccy, cycle_label, as_of_label, is_summary):
    """Letterhead block. Crimson rule + charcoal identity, and the one instruction line
    that actually matters: which cells are theirs."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(ncols, 8))
    c = ws.cell(row=1, column=1, value="CCC  |  GROUP ACCOUNTS  ·  TREASURY")
    c.font = Font(name=FONT, size=9, bold=True, color=CRIMSON)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(ncols, 8))
    title = f"Cash Flow Submission  —  {area}"
    if not is_summary:
        title += f"  ·  {project}"
    c = ws.cell(row=2, column=1, value=title)
    c.font = Font(name=FONT, size=15, bold=True, color=CHARCOAL)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=min(ncols, 10))
    sub = f"{cycle_label}   ·   Currency: {ccy}   ·   Actuals through {as_of_label}"
    c = ws.cell(row=3, column=1, value=sub)
    c.font = Font(name=FONT, size=10, color="595959")

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=min(ncols, 12))
    if is_summary:
        note = ("This sheet is calculated from the project sheets — it is locked and has "
                "no input cells. Area-level entries belong on the AREA sheet.")
    else:
        note = ("Grey cells are reported actuals, yellow cells are your forecast. "
                "Totals and balances are locked — they calculate automatically.")
    c = ws.cell(row=4, column=1, value=note)
    c.font = Font(name=FONT, size=9, italic=True, color="7F7F7F")

    for col in range(1, ncols + 1):
        ws.cell(row=5, column=col).fill = PatternFill("solid", fgColor=CRIMSON)
    ws.row_dimensions[5].height = 3


def _write_sheet(ws, *, lines, plan, cols, as_of_ym, area, project, ccy,
                 cycle_label, as_of_label, actual, forecast, summary_over=None):
    """Render one statement sheet.

    summary_over: list of sheet names -> every value cell becomes =SUM('A'!ref,'B'!ref…)
                  and the sheet carries no input cells at all.
    """
    is_summary = summary_over is not None
    ncols = 1 + len(cols)
    _style_header(ws, ncols, area, project, ccy, cycle_label, as_of_label, is_summary)

    hdr = 7
    ws.cell(row=hdr, column=1, value="Cash Flow Line").font = Font(
        name=FONT, size=10, bold=True, color=WHITE)
    ws.cell(row=hdr, column=1).fill = PatternFill("solid", fgColor=CHARCOAL)
    ws.cell(row=hdr, column=1).alignment = Alignment(vertical="center")

    month_cols, year_total_col, month_order = {}, {}, []
    for i, (kind, y, m) in enumerate(cols):
        col = 2 + i
        if kind == "m":
            label, fill, width = dt.date(y, m, 1).strftime("%b-%y"), CHARCOAL, 13
            month_cols[(y, m)] = col
            month_order.append((y, m))
        else:
            label, fill, width = f"Total {y}", CRIMSON, 15
            year_total_col[y] = col
        cell = ws.cell(row=hdr, column=col, value=label)
        cell.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.column_dimensions["A"].width = 38
    ws.row_dimensions[hdr].height = 22

    def ref(col, row):
        return f"{get_column_letter(col)}{row}"

    def is_actual(y, m):
        return y * 100 + m <= as_of_ym

    def put_value(row, col, line_code, stock=False):
        """One value cell: summary formula, year total, locked actual, or open input."""
        kind, y, m = cols[col - 2]
        cell = ws.cell(row=row, column=col)
        cell.number_format = NUM_FMT
        cell.border = BOX
        cell.font = Font(name=FONT, size=10)

        if kind == "t":
            # A stock has no meaningful annual total — never sum a level down a year.
            cell.fill = PatternFill("solid", fgColor=TOTAL_TINT)
            if stock:
                cell.value = "—"
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(name=FONT, size=10, color="BFBFBF")
            else:
                a = ref(month_cols[(y, 1)], row)
                b = ref(month_cols[(y, 12)], row)
                cell.value = f"=SUM({a}:{b})"
                cell.font = Font(name=FONT, size=10, bold=True)
            return

        if is_summary:
            parts = ",".join(f"'{s}'!{ref(col, row)}" for s in summary_over)
            cell.value = f"=SUM({parts})"
            cell.fill = PatternFill("solid", fgColor=GREY_BAND)
            return

        if is_actual(y, m):
            v = actual.get((project, line_code, y, m))
            cell.value = float(v) if v is not None else None
            cell.fill = PatternFill("solid", fgColor=GREY_BAND)
            cell.font = Font(name=FONT, size=10, color="595959")
            # Editable for now (Karim's call) — actuals stay unlocked alongside forecast.
            # Only the computed rows/columns (totals, opening/ending, nets) stay locked.
            cell.protection = Protection(locked=False)
        else:
            v = forecast.get((project, line_code, y, m))
            cell.value = float(v) if v is not None else None
            cell.fill = PatternFill("solid", fgColor=INPUT_TINT)
            cell.border = INPUT_BOX
            cell.protection = Protection(locked=False)   # <- their input

    r = hdr + 1
    line_rows = {}          # line_code -> row
    cat_net_rows = []       # rows holding each category net

    # ---- OPENING BALANCE ---------------------------------------------------------
    # Actual months carry the PUBLISHED balance verbatim (locked) — the pre-fill promise
    # is "your numbers", and deriving them would silently overwrite the plugs that exist
    # in the source data. The chain (opening = prior ending) governs the FORECAST only,
    # which is where they are doing new work and the arithmetic must hold.
    ws.cell(row=r, column=1, value="Opening Balance").font = Font(
        name=FONT, size=10, bold=True, color=CHARCOAL)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=TOTAL_TINT)
    opening_row = r
    for i, (kind, y, m) in enumerate(cols):
        col = 2 + i
        cell = ws.cell(row=r, column=col)
        cell.number_format = NUM_FMT
        cell.border = BOX
        cell.fill = PatternFill("solid", fgColor=TOTAL_TINT)
        cell.font = Font(name=FONT, size=10, bold=True)
        if kind == "t":
            cell.value = "—"
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name=FONT, size=10, color="BFBFBF")
        elif is_summary:
            parts = ",".join(f"'{s}'!{ref(col, r)}" for s in summary_over)
            cell.value = f"=SUM({parts})"
        elif is_actual(y, m):
            v = actual.get((project, OPENING, y, m))
            cell.value = float(v) if v is not None else None
            cell.protection = Protection(locked=True)
        # forecast months filled after the ending row is known (chain)
    r += 2

    # ---- FLOW SECTIONS -----------------------------------------------------------
    for kind, payload in plan:
        if kind == "band":
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            c = ws.cell(row=r, column=1, value=str(payload).upper())
            c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=CRIMSON)
            c.alignment = Alignment(vertical="center", indent=1)
            ws.row_dimensions[r].height = 18
            r += 1

        elif kind == "sub":
            c = ws.cell(row=r, column=1, value=str(payload))
            c.font = Font(name=FONT, size=9, bold=True, color="7F7F7F")
            r += 1

        elif kind == "line":
            line = payload
            c = ws.cell(row=r, column=1, value="    " + line["description"])
            c.font = Font(name=FONT, size=10, color=CHARCOAL)
            c.border = BOX
            for i in range(len(cols)):
                put_value(r, 2 + i, line["line_code"])
            line_rows[line["line_code"]] = r
            r += 1

        elif kind == "net":
            cat = payload
            cat_codes = {l["line_code"] for l in lines if l["category"] == cat}
            block = sorted(rw for lc, rw in line_rows.items() if lc in cat_codes)
            c = ws.cell(row=r, column=1, value=f"Net {cat}")
            c.font = Font(name=FONT, size=10, bold=True, color=CHARCOAL)
            c.fill = PatternFill("solid", fgColor=TOTAL_TINT)
            c.border = BOX
            for col in range(2, ncols + 1):
                cell = ws.cell(row=r, column=col)
                if block:
                    cell.value = "=" + "+".join(ref(col, rw) for rw in block)
                cell.number_format = NUM_FMT
                cell.font = Font(name=FONT, size=10, bold=True)
                cell.fill = PatternFill("solid", fgColor=TOTAL_TINT)
                cell.border = BOX
            cat_net_rows.append(r)
            r += 2

    # ---- NET MOVEMENT + ENDING (the chain that kills stale balance rows) ---------
    net_row = r
    c = ws.cell(row=r, column=1, value="NET CASH MOVEMENT")
    c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=CHARCOAL)
    for col in range(2, ncols + 1):
        cell = ws.cell(row=r, column=col)
        if cat_net_rows:
            cell.value = "=" + "+".join(ref(col, rw) for rw in cat_net_rows)
        cell.number_format = NUM_FMT
        cell.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=CHARCOAL)
    r += 1

    ending_row = r
    c = ws.cell(row=r, column=1, value="Ending Balance")
    c.font = Font(name=FONT, size=10, bold=True, color=CHARCOAL)
    c.fill = PatternFill("solid", fgColor=TOTAL_TINT)
    c.border = BOX
    for i, (kind, y, m) in enumerate(cols):
        col = 2 + i
        cell = ws.cell(row=r, column=col)
        cell.number_format = NUM_FMT
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.fill = PatternFill("solid", fgColor=TOTAL_TINT)
        cell.border = BOX
        if kind == "t":
            cell.value = "—"
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name=FONT, size=10, color="BFBFBF")
        elif is_summary:
            parts = ",".join(f"'{s}'!{ref(col, r)}" for s in summary_over)
            cell.value = f"=SUM({parts})"
        elif is_actual(y, m):
            v = actual.get((project, ENDING, y, m))
            cell.value = float(v) if v is not None else None
            cell.protection = Protection(locked=True)
        else:
            cell.value = f"={ref(col, opening_row)}+{ref(col, net_row)}"

    # Chain the FORECAST region only: each forecast month opens on the prior month's close.
    if not is_summary:
        for idx, (y, m) in enumerate(month_order):
            if idx == 0 or is_actual(y, m):
                continue
            py, pm = month_order[idx - 1]
            ws.cell(row=opening_row, column=month_cols[(y, m)]).value = \
                f"={ref(month_cols[(py, pm)], ending_row)}"
    r += 2

    # ---- STOCK MEMOS (loans / overdrafts — levels, never summed) -----------------
    ws.cell(row=r, column=1, value="MEMO — CLOSING STOCKS").font = Font(
        name=FONT, size=9, bold=True, color="7F7F7F")
    r += 1
    for code in MEMO_STOCKS:
        line = next((l for l in lines if l["line_code"] == code), None)
        if not line:
            continue
        c = ws.cell(row=r, column=1, value="    " + line["description"])
        c.font = Font(name=FONT, size=10, color=CHARCOAL)
        c.border = BOX
        for i in range(len(cols)):
            put_value(r, 2 + i, code, stock=True)
        r += 1

    ws.freeze_panes = ws.cell(row=hdr + 1, column=2)
    ws.sheet_view.showGridLines = False

    # Protection — every sheet protected, every cell SELECTABLE (read the formula), and:
    #   SUMMARY  — nothing unlocked -> fully read-only.
    #   Projects — the VALUE cells are unlocked (actuals + forecast, editable for now);
    #              the computed rows/columns stay locked: year totals, opening & ending
    #              balance, and the Net rows are all formulas that were never unlocked.
    # OOXML gotcha that caused "I can't even select a cell": selectLockedCells="1" /
    # selectUnlockedCells="1" DISABLE selection. To allow selecting, they must be False
    # (the default). Do NOT set them True.
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False    # allow selecting locked cells
    ws.protection.selectUnlockedCells = False  # allow selecting unlocked cells
    ws.protection.formatCells = False

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _write_meta(wb, *, area, version, cycle_label, as_of_ym, ccy, sheet_map, years):
    """Hidden contract sheet. This is what makes intake mechanical instead of guesswork."""
    ws = wb.create_sheet("_meta")
    rows = [
        ("template_version", TEMPLATE_VERSION),
        ("generated_at", dt.datetime.utcnow().isoformat(timespec="seconds") + "Z"),
        ("area", area),
        ("source_version", version),
        ("cycle_label", cycle_label),
        ("as_of_ym", as_of_ym),
        ("currency", ccy),
        ("scale", 1),          # values are FULL native units - no scale factor, by design
        ("years", ",".join(str(y) for y in years)),
    ]
    for i, (k, v) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=k).font = Font(name=FONT, size=9, bold=True)
        ws.cell(row=i, column=2, value=v).font = Font(name=FONT, size=9)
    r = len(rows) + 2
    ws.cell(row=r, column=1, value="sheet_name").font = Font(name=FONT, size=9, bold=True)
    ws.cell(row=r, column=2, value="project_code").font = Font(name=FONT, size=9, bold=True)
    for sheet, code in sheet_map.items():
        r += 1
        ws.cell(row=r, column=1, value=sheet)
        ws.cell(row=r, column=2, value=code)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.sheet_state = "hidden"
    ws.protection.sheet = True


def _safe(name):
    """Excel sheet-name constraints: 31 chars, no []:*?/\\"""
    out = "".join("-" if ch in "[]:*?/\\" else ch for ch in str(name))
    return out[:31] or "Sheet"


# =================================================================================
# Entry point
# =================================================================================
def build_area_template(area, version, as_of_ym, cycle_label=None):
    """Return (filename, xlsx_bytes) for one area.

    as_of_ym: int YYYYMM — the last ACTUAL month. Everything after it is their input.
    """
    cycle_year = as_of_ym // 100
    as_of_label = dt.date(cycle_year, as_of_ym % 100, 1).strftime("%B %Y")
    cycle_label = cycle_label or version

    lines = load_lines()
    actual, forecast, ccy, years, projects = load_cells(area, version, cycle_year, as_of_ym)
    if not projects:
        raise ValueError(f"No cash-flow data for area '{area}' in version '{version}'")

    # The AREA sheet (the area-level rollup) leads the project sheets, right after SUMMARY;
    # everything else follows alphabetically. Codes vary: 'AREA' or '_AREA'.
    def _is_area(code):
        return str(code).strip().upper().lstrip("_").strip() == "AREA"
    projects = sorted(projects, key=lambda c: (0 if _is_area(c) else 1, str(c)))

    cols = _col_plan(years)
    plan = _row_plan(lines)

    wb = Workbook()
    wb.remove(wb.active)

    sheet_map = {}
    for code in projects:
        name = _safe(code)
        ws = wb.create_sheet(name)
        _write_sheet(ws, lines=lines, plan=plan, cols=cols, as_of_ym=as_of_ym,
                     area=area, project=code, ccy=ccy, cycle_label=cycle_label,
                     as_of_label=as_of_label, actual=actual, forecast=forecast)
        sheet_map[name] = code

    # SUMMARY last in build order, first in tab order — pure formulas over the projects.
    if len(projects) > 1:
        ws = wb.create_sheet("SUMMARY")
        _write_sheet(ws, lines=lines, plan=plan, cols=cols, as_of_ym=as_of_ym,
                     area=area, project="SUMMARY", ccy=ccy, cycle_label=cycle_label,
                     as_of_label=as_of_label, actual=actual, forecast=forecast,
                     summary_over=list(sheet_map.keys()))
        wb.move_sheet("SUMMARY", offset=-len(wb.sheetnames) + 1)
        sheet_map["SUMMARY"] = "_SUMMARY"

    _write_meta(wb, area=area, version=version, cycle_label=cycle_label,
                as_of_ym=as_of_ym, ccy=ccy, sheet_map=sheet_map, years=years)

    # Open on SUMMARY (index 0 for a multi-project area; the single sheet otherwise).
    # SUMMARY is locked but selectable now, so landing here no longer reads as "locked".
    wb.active = 0
    for i, ws in enumerate(wb.worksheets):
        ws.sheet_view.tabSelected = (i == 0)

    buf = io.BytesIO()
    wb.save(buf)
    slug = "".join(ch if ch.isalnum() else "-" for ch in area).strip("-")
    return f"CashFlow-{slug}-{as_of_ym}.xlsx", buf.getvalue()


def list_areas(version):
    """Areas available to generate for a version (for the UI picker)."""
    rows = db.select("cf_forecasts", {"select": "area", "version": f"eq.{version}"})
    return sorted({r["area"] for r in rows})
